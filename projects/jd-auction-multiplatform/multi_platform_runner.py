
import argparse
import json
import os
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Mapping, Optional
try:
    from typing import Protocol, Optional
except ImportError:
    class Protocol:
        pass
from urllib.parse import urljoin

import requests
import threading

from jd.logger import get_logger

from jd.ai_config import load_default_dotenv

load_default_dotenv()

logger = get_logger()

import jd_scraper_v2 as jd_v2
from jd.ai_extractor import AIExtractionContext
from jd_mysql_store import MySQLConfig, MySQLJDScraperDatabase, mysql_connection, reset_mysql_tables
from platform_adapters.ali_adapter import (
    ALI_DEFAULT_DETAIL_URL,
    ALI_LIST_CHANNELS,
    ALI_SOURCE_PLATFORM,
    AliAuctionAdapter,
    AliDetailBundle,
    AliListItem,
)
from platform_adapters.cquae_adapter import (
    CQUAE_BASE_URL,
    CQUAE_DATA_SOURCE,
    CQUAE_PLATFORM,
    CquaeAdapter,
    CquaeBrowserFetcher,
    CquaeDetailBundle,
    CquaeListItem,
)
from platform_adapters.sdcqjy_adapter import (
    SDCQJY_DATA_SOURCE,
    SDCQJY_LIST_ENDPOINT,
    SDCQJY_PLATFORM,
    SdcqjyAdapter,
)
from platform_adapters.ejy365_adapter import (
    DEFAULT_BASE_URL as EJY365_BASE_URL,
    Ejy365Adapter,
    Ejy365DetailBundle,
    Ejy365ListItem,
)
from platform_adapters.tpre_adapter import (
    TPRE_BASE_URL,
    TPRE_DATA_SOURCE,
    TPRE_PLATFORM,
    TpreAdapter,
    TpreDetailBundle,
    TpreListItem,
)
from platform_adapters.prechina_adapter import (
    PRECHINA_BASE_URL,
    PRECHINA_DATA_SOURCE,
    PRECHINA_PLATFORM,
    PrechinaAdapter,
    PrechinaDetailBundle,
    PrechinaListItem,
)
from platform_adapters.gxcq_adapter import (
    GXCQ_BASE_URL,
    GXCQ_DATA_SOURCE,
    GXCQ_PLATFORM,
    GxcqAdapter,
    GxcqDetailBundle,
    GxcqListItem,
)
from platform_adapters.gycq_adapter import (
    GYCQ_DATA_SOURCE,
    GYCQ_PLATFORM,
    GycqAdapter,
)
from platform_adapters.cbex_adapter import (
    CBEX_DATA_SOURCE,
    CBEX_PLATFORM,
    CbexAdapter,
    CbexBrowserFetcher,
    CbexDetailBundle,
    CbexListItem,
    OTCPRJ_DETAIL_URL,
)
from platform_adapters.jd_adapter import (
    JD_DATA_SOURCE,
    JD_SOURCE_PLATFORM,
    JDPlatformAdapter,
)


DEFAULT_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/126.0 Safari/537.36"
    ),
    "Accept-Language": "zh-CN,zh;q=0.9",
}

AI_UNVERIFIED_COMMON_FIELDS = {
    "attachments_json",
}

AI_FILL_ONLY_COMMON_FIELDS = {
    "asset_type",
    "project_name",
    "data_source",
    "source_platform",
    "source_item_id",
    "source_url",
    "source_site_name",
    "start_price_raw",
    "final_price_raw",
    "bid_records_json",
}


EJY365_PROJECT_TYPE_LABELS: dict[str, tuple[str, str]] = {
    "ZQ": ("debt", "债权"),
    "FC": ("real_estate", "房地产"),
    "FCZZ": ("real_estate", "房产租赁"),
    "TD": ("land", "土地"),
    "CL": ("vehicle", "车辆"),
    "GQ": ("equity", "股权"),
    "ZSCQ": ("ip", "知识产权"),
    "WZ": ("goods", "物资产品"),
    "ZYSYQ": ("usufruct", "用益物权"),
    "KQ": ("usufruct", "矿权"),
    "LQ": ("usufruct", "林权"),
    "HKQ": ("usufruct", "海域使用权"),
    "QT": ("other", "其他"),
}
DEFAULT_EJY365_PROJECT_TYPES = (
    "ZQ", "FC", "FCZZ", "CL", "GQ", "TD", "ZSCQ", "WZ", "ZYSYQ",
    "KQ", "LQ", "HKQ", "GGJYQ", "ZJGC", "MTQCNZB", "PFQJY", "STZY",
    "KCPZR", "QYTPH", "SJCPL", "SLL", "SB", "CZ", "GYYSQ", "TDFWQ",
    "QTKQ", "TZQ", "JYQ", "CSSYQ", "ZRZY", "PWS", "PWFQ", "HJSYQ",
    "JN", "SJ", "QT",
)


def ejy365_asset_for_project_type(project_type: Any) -> tuple[str, str]:
    code = compact_text(project_type).upper()
    return EJY365_PROJECT_TYPE_LABELS.get(code, ("other", "其他"))


@dataclass
class PlatformRecord:
    source_platform: str
    source_site_name: str
    source_item_id: str
    source_url: str
    asset_group: str
    category_id: str = ""
    category_name: str = ""
    common_values: dict[str, Any] = field(default_factory=dict)
    field_results: dict[str, dict[str, Any]] = field(default_factory=dict)
    special_values: dict[str, Any] = field(default_factory=dict)
    special_field_results: dict[str, dict[str, Any]] = field(default_factory=dict)
    raw_payloads: dict[str, Any] = field(default_factory=dict)
    attachments_json: Any = None
    debt_details: list[dict[str, Any]] = field(default_factory=list)
    ip_details: list[dict[str, Any]] = field(default_factory=list)


@dataclass
class PlatformCrawlResult:
    platform: str
    batch_id: str
    scanned_count: int = 0
    success_count: int = 0
    failed_count: int = 0
    errors: list[dict[str, str]] = field(default_factory=list)
    list_total: int = 0          # 鍒楄〃椤靛疄闄呰繑鍥炵殑鎬绘潯鏁?澧為噺妯″紡涓嬩负鍏ㄩ噺鍒楄〃闀垮害)
    skipped_existing: int = 0    # 澧為噺妯″紡涓嬪洜"宸查噰闆嗕笖鍒楄〃鎸囩汗鏈彉"璺宠繃鐨勬潯鏁?


class PlatformHandler(Protocol):
    source_platform: str
    source_site_name: str

    def fetch_list(self, limit: int) -> list[Any]:
        ...

    def fetch_detail(self, list_item: Any) -> Any:
        ...

    def build_record(self, detail_bundle: Any) -> PlatformRecord:
        ...


def _list_item_id(item: Any) -> str:
    """浠庡钩鍙板垪琛ㄩ」涓彁鍙栫ǔ瀹氱殑鍞竴 ID, 鍏煎鍚勫钩鍙颁笉鍚岀殑 id 瀛楁鍚嶃?

    浼樺厛绾? source_item_id > prj_id > item_id > project_no > slug > detail_url > source_url
    鐢ㄤ簬澧為噺鍘婚噸涓庡垪琛ㄦ寚绾? 纭繚 Cbex/Ali/Ejy365 绛夊垪琛ㄩ」缂哄皯 source_item_id 瀛楁鏃?
    涔熻兘姝ｇ‘鎻愬彇 ID(鍚﹀垯澧為噺鍘婚噸浼氶鍖栦负鍙繚鐣?1 鏉?銆?
    """
    keys = (
        "source_item_id",
        "prj_id",
        "item_id",
        "project_no",
        "slug",
        "id",
        "detail_url",
        "source_url",
        "url",
    )
    if isinstance(item, Mapping):
        for key in keys:
            text = compact_text(item.get(key))
            if text:
                return text
    return compact_text(
        getattr(item, "source_item_id", None)
        or getattr(item, "prj_id", None)
        or getattr(item, "item_id", None)
        or getattr(item, "project_no", None)
        or getattr(item, "slug", None)
        or getattr(item, "id", None)
        or getattr(item, "detail_url", None)
        or getattr(item, "source_url", None)
        or getattr(item, "url", None)
    ) or ""


def _checkpoint_value(checkpoint: Any, key: str) -> Any:
    if not checkpoint:
        return None
    if isinstance(checkpoint, Mapping):
        return checkpoint.get(key)
    return getattr(checkpoint, key, None)


def _checkpoint_status(checkpoint: Any) -> str:
    return compact_text(_checkpoint_value(checkpoint, "checkpoint_status") or "running").lower()


def _should_resume_from_checkpoint(mode: str, checkpoint: Any) -> bool:
    normalized_mode = compact_text(mode).lower()
    if normalized_mode not in {"full", "incremental"} or not checkpoint:
        return False
    if not compact_text(_checkpoint_value(checkpoint, "last_item_id")):
        return False
    return _checkpoint_status(checkpoint) not in {"completed", "success"}


def _items_after_checkpoint(items: list[Any], checkpoint: Any) -> list[Any]:
    last_item_id = compact_text(_checkpoint_value(checkpoint, "last_item_id"))
    if not last_item_id:
        return list(items)
    found = False
    remaining: list[Any] = []
    for item in items:
        if found:
            remaining.append(item)
        elif _list_item_id(item) == last_item_id:
            found = True
    return remaining if found else list(items)


def request_timeout_value(timeout: int | float | None) -> int | float | None:
    if timeout is None:
        return None
    try:
        numeric = float(timeout)
    except (TypeError, ValueError):
        return timeout
    return None if numeric <= 0 else timeout


def compact_text(value: Any) -> str:
    return jd_v2.compact_text(value)


def _parse_task_types_arg(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, str):
        raw_items = [part.strip() for part in re.split(r"[,/;锛屻乗s]+", value) if part.strip()]
    elif isinstance(value, (list, tuple, set)):
        raw_items = [str(part).strip() for part in value if str(part).strip()]
    else:
        raw_items = [str(value).strip()]
    result: list[str] = []
    for item in raw_items:
        if item and item not in result:
            result.append(item)
    return result


def safe_json(value: Any) -> str:
    return jd_v2.safe_json_dumps(value)


def make_field_result(
    value: Any,
    source_payload_type: str,
    source_path: str,
    excerpt: Any = None,
    *,
    method: str = "html_rule",
    confidence: Optional[float] = None,
) -> dict[str, Any]:
    return jd_v2.field_result_value(
        value,
        source_payload_type,
        source_path,
        compact_text(excerpt) if excerpt is not None else None,
        method=method,
        confidence=confidence,
    )


def normalize_attachments_payload(files: Any = None, media: Any = None) -> dict[str, Any]:
    def parse_payload(value: Any) -> Any:
        if isinstance(value, str):
            try:
                return json.loads(value)
            except json.JSONDecodeError:
                return value
        return value

    def iter_values(value: Any) -> Iterable[Any]:
        value = parse_payload(value)
        if isinstance(value, Mapping):
            yield value
            for item in value.values():
                yield from iter_values(item)
        elif isinstance(value, list):
            for item in value:
                yield from iter_values(item)

    def clean_url(value: Any) -> str:
        url = compact_text(value)
        if not url:
            return ""
        if url.lower().startswith(("javascript:", "about:", "#")):
            return ""
        return url

    def file_url(entry: Mapping[str, Any]) -> str:
        for key in (
            "url",
            "href",
            "attachmentAddress",
            "fileUrl",
            "downloadUrl",
            "downloadURL",
            "attachmentUrl",
            "filePath",
            "path",
            "src",
            "resourceUrl",
            "resourceURL",
            "ossUrl",
            "previewUrl",
        ):
            url = clean_url(entry.get(key))
            if url:
                return url
        return ""

    def normalize_file(entry: Any) -> dict[str, Any] | None:
        if isinstance(entry, str):
            url = clean_url(entry)
            if not url.startswith(("http://", "https://", "//")):
                return None
            return {"name": url.rsplit("/", 1)[-1], "url": url}
        if not isinstance(entry, Mapping):
            return None
        url = file_url(entry)
        if not url:
            return None
        normalized = dict(entry)
        normalized["url"] = url
        if not compact_text(normalized.get("name")):
            normalized["name"] = jd_v2.first_non_blank(
                normalized.get("attachmentName"),
                normalized.get("fileName"),
                normalized.get("file_name"),
                normalized.get("title"),
                normalized.get("label"),
                normalized.get("attachmentCode"),
            ) or url.rsplit("/", 1)[-1]
        return normalized

    def contains_media_marker(value: Any) -> bool:
        value = parse_payload(value)
        if isinstance(value, Mapping):
            for key, item in value.items():
                low = str(key).lower()
                if low in {
                    "media",
                    "imagevideoarea",
                    "imagelist",
                    "videolist",
                    "imagepath",
                    "imageurl",
                    "imgurl",
                    "picurl",
                    "videopath",
                    "videourl",
                }:
                    return True
                if contains_media_marker(item):
                    return True
        elif isinstance(value, list):
            return any(contains_media_marker(item) for item in value)
        return False

    def contains_media_reference(value: Any) -> bool:
        value = parse_payload(value)
        if isinstance(value, Mapping):
            for key, item in value.items():
                low = str(key).lower()
                if low in {
                    "imagepath",
                    "imageurl",
                    "imgurl",
                    "picurl",
                    "src",
                    "videopath",
                    "videourl",
                    "video",
                } and compact_text(item):
                    return True
                if contains_media_reference(item):
                    return True
        elif isinstance(value, list):
            return any(contains_media_reference(item) for item in value)
        return False

    files_value = parse_payload(files)
    file_candidates: list[Any] = []
    if isinstance(files_value, Mapping):
        for key in ("files", "data", "attachments", "attachmentList", "attachList", "attachFiles", "fileList", "docs", "documents"):
            item = files_value.get(key)
            if item is not None:
                file_candidates.append(item)
        file_candidates.append(files_value)
    elif isinstance(files_value, list):
        file_candidates.append(files_value)
    elif files_value:
        file_candidates.append(files_value)

    normalized_files: list[dict[str, Any]] = []
    seen_file_urls: set[str] = set()
    for candidate in file_candidates:
        for entry in iter_values(candidate):
            normalized = normalize_file(entry)
            if not normalized:
                continue
            url = normalized["url"]
            if url in seen_file_urls:
                continue
            seen_file_urls.add(url)
            normalized_files.append(normalized)
    payload: dict[str, Any] = {"files": normalized_files}

    media_values: list[Any] = []
    if isinstance(files_value, Mapping) and files_value.get("media") is not None and contains_media_reference(files_value.get("media")):
        media_values.append(files_value.get("media"))
    if files_value is not None and contains_media_marker(files_value) and contains_media_reference(files_value):
        media_values.append(files_value)
    media_value = parse_payload(media)
    if media_value and contains_media_reference(media_value):
        media_values.append(media_value)
    payload["media"] = media_values
    return payload


def common_results_from_values(values: Mapping[str, Any], source_type: str, source_path: str) -> dict[str, dict[str, Any]]:
    results: dict[str, dict[str, Any]] = {}
    common_keys = {field.key for field in jd_v2.COMMON_FIELDS}
    for key, value in values.items():
        if key not in common_keys:
            continue
        results[key] = make_field_result(value, source_type, f"{source_path}.{key}", value)
    return results


def special_results_from_values(
    asset_group: str,
    values: Mapping[str, Any],
    source_type: str,
    source_path: str,
) -> dict[str, dict[str, Any]]:
    results: dict[str, dict[str, Any]] = {}
    special_keys = {field.key for field in jd_v2.SPECIAL_FIELDS.get(asset_group, ())}
    for key, value in values.items():
        if key not in special_keys:
            continue
        results[key] = make_field_result(value, source_type, f"{source_path}.{key}", value)
    return results


def split_ai_results(
    ai_results: Mapping[str, Any],
    asset_group: str,
) -> tuple[dict[str, Any], dict[str, dict[str, Any]], dict[str, Any], dict[str, dict[str, Any]]]:
    common_keys = {field.key for field in jd_v2.COMMON_FIELDS}
    special_keys = {field.key for field in jd_v2.SPECIAL_FIELDS.get(asset_group, ())}
    if asset_group == "other":
        special_keys.add("extracted_summary")

    common_values: dict[str, Any] = {}
    common_results: dict[str, dict[str, Any]] = {}
    special_values: dict[str, Any] = {}
    special_results: dict[str, dict[str, Any]] = {}
    for key, result in ai_results.items():
        if key in AI_UNVERIFIED_COMMON_FIELDS:
            continue
        value = getattr(result, "value", None)
        if jd_v2.is_blank(value):
            continue
        field_result = make_field_result(
            value,
            "ai_extraction",
            "llm_batch",
            getattr(result, "original_text", "") or getattr(result, "reasoning", "") or value,
            method="ai",
            confidence=getattr(result, "confidence", 0.75),
        )
        if key in common_keys:
            common_values[key] = value
            common_results[key] = field_result
        elif key in special_keys:
            special_values[key] = value
            special_results[key] = field_result
    return common_values, common_results, special_values, special_results


def ai_result_to_field_result(result: Any, *, source_path: str, method: str = "ai") -> dict[str, Any]:
    value = getattr(result, "value", None)
    excerpt = getattr(result, "original_text", "") or getattr(result, "reasoning", "") or value
    return make_field_result(
        value,
        "ai_extraction",
        source_path,
        excerpt,
        method=method,
        confidence=getattr(result, "confidence", 0.75),
    )


def ai_context_to_payload(context: AIExtractionContext) -> dict[str, Any]:
    return asdict(context)


def ai_context_from_payload(payload: Any) -> AIExtractionContext:
    if isinstance(payload, AIExtractionContext):
        return payload
    if isinstance(payload, str):
        try:
            payload = json.loads(payload)
        except json.JSONDecodeError:
            payload = {}
    if not isinstance(payload, dict):
        payload = {}
    return AIExtractionContext(
        html_key_values=payload.get("html_key_values") or {},
        detail_text=payload.get("detail_text") or "",
        notice_text=payload.get("notice_text") or "",
        image_urls=payload.get("image_urls") or [],
        asset_group=payload.get("asset_group") or "",
        paimai_id=payload.get("paimai_id") or "",
    )


def ai_fields_for_asset_group(asset_group: str) -> list[tuple[str, str, str]]:
    fields = [jd_v2.ai_field_tuple(field) for field in jd_v2.COMMON_FIELDS]
    fields.extend(jd_v2.special_ai_field_tuples(asset_group))
    return fields


def ai_results_to_payload(ai_results: Mapping[str, Any]) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    for key, result in ai_results.items():
        if hasattr(result, "__dataclass_fields__"):
            payload[key] = asdict(result)
        else:
            payload[key] = result
    return payload


class MultiPlatformRunner:
    def __init__(
        self,
        db: Any,
        handlers: Mapping[str, PlatformHandler],
        *,
        ai_enabled: bool = True,
        ai_mode: str = "sync",
        item_concurrency: int = 1,
        parse_attachments: bool = False,
    ) -> None:
        self.db = db
        self.handlers = dict(handlers)
        normalized_mode = (ai_mode or "sync").strip().lower()
        if normalized_mode not in {"sync", "async", "off"}:
            raise ValueError("ai_mode must be one of: sync, async, off")
        self.ai_mode = "off" if not ai_enabled else normalized_mode
        self.ai_enabled = self.ai_mode != "off"
        self.item_concurrency = max(1, int(item_concurrency or 1))
        self.parse_attachments = parse_attachments

    def _load_crawl_checkpoint(self, platform: str, category_key: str = "default") -> Any:
        loader = getattr(self.db, "load_checkpoint", None)
        if not callable(loader):
            return None
        try:
            return loader(source_platform=platform, category_key=category_key)
        except Exception:
            return None

    def _save_crawl_checkpoint(
        self,
        *,
        platform: str,
        batch_id: str,
        mode: str,
        status: str,
        category_key: str = "default",
        current_page: int = 1,
        total_items_seen: int = 0,
        last_item_id: str | None = None,
        message: str | None = None,
    ) -> None:
        saver = getattr(self.db, "save_checkpoint", None)
        if not callable(saver):
            return
        try:
            completed_at = jd_v2.now_text() if status in {"completed", "success", "failed"} else None
            saver(
                source_platform=platform,
                category_key=category_key,
                current_page=current_page,
                total_items_seen=max(0, int(total_items_seen or 0)),
                last_item_id=last_item_id,
                batch_id=batch_id,
                crawl_mode=mode,
                checkpoint_status=status,
                message=message,
                completed_at=completed_at,
            )
        except Exception:
            pass

    def crawl_platform(self, platform: str, limit: int = 10, mode: str = "sample") -> PlatformCrawlResult:
        if platform not in self.handlers:
            raise KeyError(f"unknown platform: {platform}")

        handler = self.handlers[platform]
        normalized_mode = compact_text(mode).lower() or "sample"
        category_key = "default"
        eff_concurrency = 1 if platform == "cbex" else self.item_concurrency
        batch_id = f"{platform}-{int(time.time())}"
        result = PlatformCrawlResult(platform=platform, batch_id=batch_id)

        checkpoint = self._load_crawl_checkpoint(platform, category_key)
        resume_enabled = _should_resume_from_checkpoint(normalized_mode, checkpoint)
        checkpoint_seen = int(_checkpoint_value(checkpoint, "total_items_seen") or 0) if resume_enabled else 0
        checkpoint_last_id = compact_text(_checkpoint_value(checkpoint, "last_item_id")) if resume_enabled else ""
        last_processed_id: str | None = checkpoint_last_id or None
        processed_since_start = 0
        status = "failed"

        crawl_with_db = getattr(handler, "crawl_with_db", None)
        if callable(crawl_with_db):
            def adapter_checkpoint(**progress: Any) -> None:
                self._save_crawl_checkpoint(
                    platform=platform,
                    batch_id=compact_text(progress.get("batch_id")) or batch_id,
                    mode=normalized_mode,
                    status=compact_text(progress.get("status")) or "running",
                    category_key=compact_text(progress.get("category_key")) or category_key,
                    current_page=int(progress.get("current_page") or 1),
                    total_items_seen=int(progress.get("total_items_seen") or 0),
                    last_item_id=compact_text(progress.get("last_item_id")) or None,
                    message=compact_text(progress.get("message")) or "adapter crawl progress",
                )

            self._save_crawl_checkpoint(
                platform=platform, batch_id=batch_id, mode=normalized_mode, status="running",
                category_key=category_key, total_items_seen=checkpoint_seen,
                last_item_id=last_processed_id, message="adapter crawl_with_db started",
            )
            try:
                result = crawl_with_db(
                    self.db,
                    limit,
                    normalized_mode,
                    ai_mode=self.ai_mode,
                    checkpoint_callback=adapter_checkpoint,
                    resume_checkpoint=checkpoint if resume_enabled else None,
                )
                if not getattr(result, "batch_id", None):
                    result.batch_id = batch_id
                seen = max(checkpoint_seen, int(getattr(result, "success_count", 0) or 0))
                self._save_crawl_checkpoint(
                    platform=platform, batch_id=result.batch_id, mode=normalized_mode, status="completed",
                    category_key=category_key, total_items_seen=seen,
                    last_item_id=last_processed_id, message="adapter crawl_with_db completed",
                )
                return result
            except Exception as exc:
                result.errors.append({"item": "adapter", "error": str(exc)})
                self._save_crawl_checkpoint(
                    platform=platform, batch_id=batch_id, mode=normalized_mode, status="failed",
                    category_key=category_key, total_items_seen=checkpoint_seen,
                    last_item_id=last_processed_id, message=str(exc),
                )
                return result
            finally:
                close_handler = getattr(handler, "close", None)
                if callable(close_handler):
                    try:
                        close_handler()
                    except Exception:
                        pass

        batch_id = self.db.start_batch(
            {
                "source_platform": platform,
                "source_site_name": getattr(handler, "source_site_name", platform),
                "limit": limit,
                "mode": normalized_mode,
                "runner": "multi_platform_runner",
            }
        )
        result.batch_id = batch_id
        full_list_for_fp: list[Any] = []
        list_items: list[Any] = []
        streamed_list = False

        try:
            self._save_crawl_checkpoint(
                platform=platform, batch_id=batch_id, mode=normalized_mode, status="running",
                category_key=category_key, total_items_seen=checkpoint_seen,
                last_item_id=last_processed_id,
                message="crawl started" if not resume_enabled else "crawl resumed from checkpoint",
            )

            if normalized_mode == "incremental":
                existing_ids = self._query_existing_ids(platform)
                known_fps = self._query_list_fingerprints(platform)
                has_baseline = bool(known_fps)
                fp_method = getattr(handler, "list_fingerprint", None)

                if checkpoint and has_baseline:
                    try:
                        quick_list = handler.fetch_list(100)
                        quick_new = [i for i in quick_list if _list_item_id(i) not in existing_ids]
                        if not quick_new and quick_list:
                            result.skipped_existing = len(quick_list)
                            result.list_total = 0
                            result.scanned_count = 0
                            result.success_count = 0
                            status = "success"
                            self._save_crawl_checkpoint(
                                platform=platform, batch_id=batch_id, mode=normalized_mode, status="completed",
                                category_key=category_key, total_items_seen=checkpoint_seen,
                                last_item_id=last_processed_id,
                                message="incremental quick scan found no new items",
                            )
                            self.db.finish_batch(batch_id, "success", safe_json(result.__dict__))
                            return result
                    except Exception:
                        pass

                full_list = handler.fetch_list(0)
                deduped: dict[str, Any] = {}
                for item in full_list:
                    sid = _list_item_id(item)
                    if sid and sid not in deduped:
                        deduped[sid] = item
                full_list_for_fp = list(deduped.values())
                result.list_total = len(full_list)
                to_crawl: list[Any] = []
                skipped = 0
                consecutive_old = 0
                stop_after_consecutive_old = 50
                for item in full_list_for_fp:
                    sid = _list_item_id(item)
                    if sid not in existing_ids:
                        to_crawl.append(item)
                        consecutive_old = 0
                        continue
                    if callable(fp_method):
                        cur_fp = fp_method(item)
                        known = known_fps.get(sid)
                        if known is not None and known == cur_fp:
                            skipped += 1
                            consecutive_old += 1
                            if consecutive_old >= stop_after_consecutive_old:
                                break
                            continue
                        if known is None and not has_baseline:
                            skipped += 1
                            consecutive_old += 1
                            if consecutive_old >= stop_after_consecutive_old:
                                break
                            continue
                    else:
                        skipped += 1
                        consecutive_old += 1
                        if consecutive_old >= stop_after_consecutive_old:
                            break
                        continue
                    to_crawl.append(item)
                    consecutive_old = 0
                result.skipped_existing = skipped
                list_items = _items_after_checkpoint(to_crawl, checkpoint) if resume_enabled else to_crawl
                if resume_enabled:
                    result.skipped_existing += checkpoint_seen
            else:
                iter_list_pages = getattr(handler, "iter_list_pages", None)
                if callable(iter_list_pages):
                    streamed_list = True
                    stream_limit = 0 if normalized_mode == "full" else limit
                    checkpoint_found = not resume_enabled
                    if resume_enabled:
                        result.skipped_existing = checkpoint_seen
                    stop_stream = False
                    for page_no, page_items in enumerate(iter_list_pages(stream_limit), start=1):
                        page_items = list(page_items or [])
                        if not page_items:
                            continue
                        full_list_for_fp.extend(page_items)
                        result.list_total += len(page_items)
                        self._save_crawl_checkpoint(
                            platform=platform, batch_id=batch_id, mode=normalized_mode, status="running",
                            category_key=category_key, current_page=page_no,
                            total_items_seen=checkpoint_seen + processed_since_start,
                            last_item_id=last_processed_id, message="streaming list page",
                        )
                        for list_item in page_items:
                            item_id = _list_item_id(list_item)
                            if resume_enabled and not checkpoint_found:
                                if item_id == checkpoint_last_id:
                                    checkpoint_found = True
                                continue
                            self._crawl_list_item(handler, batch_id, list_item, result)
                            processed_since_start += 1
                            result.scanned_count += 1
                            last_processed_id = item_id or last_processed_id
                            self._save_crawl_checkpoint(
                                platform=platform, batch_id=batch_id, mode=normalized_mode, status="running",
                                category_key=category_key, current_page=page_no,
                                total_items_seen=checkpoint_seen + processed_since_start,
                                last_item_id=last_processed_id, message="crawl progress",
                            )
                            if stream_limit and result.scanned_count >= stream_limit:
                                stop_stream = True
                                break
                        if stop_stream:
                            break
                    list_items = []
                else:
                    list_items = handler.fetch_list(0 if normalized_mode == "full" else limit)
                    full_list_for_fp = list(list_items)
                    result.list_total = len(list_items)
                    if resume_enabled:
                        list_items = _items_after_checkpoint(list_items, checkpoint)
                        result.skipped_existing = checkpoint_seen

            if not streamed_list:
                result.scanned_count = len(list_items)
            if eff_concurrency <= 1 or len(list_items) <= 1:
                for list_item in list_items:
                    self._crawl_list_item(handler, batch_id, list_item, result)
                    processed_since_start += 1
                    last_processed_id = _list_item_id(list_item) or last_processed_id
                    self._save_crawl_checkpoint(
                        platform=platform, batch_id=batch_id, mode=normalized_mode, status="running",
                        category_key=category_key, total_items_seen=checkpoint_seen + processed_since_start,
                        last_item_id=last_processed_id, message="crawl progress",
                    )
            else:
                with ThreadPoolExecutor(max_workers=min(eff_concurrency, len(list_items))) as executor:
                    futures = {
                        executor.submit(self._crawl_list_item, handler, batch_id, list_item, None): list_item
                        for list_item in list_items
                    }
                    for future in as_completed(futures):
                        list_item = futures[future]
                        try:
                            future.result()
                            result.success_count += 1
                            processed_since_start += 1
                            last_processed_id = _list_item_id(list_item) or last_processed_id
                            self._save_crawl_checkpoint(
                                platform=platform, batch_id=batch_id, mode=normalized_mode, status="running",
                                category_key=category_key, total_items_seen=checkpoint_seen + processed_since_start,
                                last_item_id=last_processed_id, message="crawl progress",
                            )
                        except Exception as exc:
                            result.failed_count += 1
                            result.errors.append({
                                "item": compact_text(getattr(list_item, "source_item_id", "")) or compact_text(list_item),
                                "error": str(exc),
                            })

            status = "success" if result.failed_count == 0 else ("partial_success" if result.success_count else "failed")
            if full_list_for_fp and normalized_mode in ("incremental", "full", "sample"):
                fp_method = getattr(handler, "list_fingerprint", None)
                if callable(fp_method):
                    self._update_list_fingerprints(platform, full_list_for_fp, handler)
            return result
        except Exception as exc:
            result.failed_count = max(result.failed_count, limit)
            result.errors.append({"item": "list", "error": str(exc)})
            return result
        finally:
            checkpoint_status = "completed" if status in {"success", "partial_success"} else "failed"
            self._save_crawl_checkpoint(
                platform=platform, batch_id=batch_id, mode=normalized_mode, status=checkpoint_status,
                category_key=category_key, total_items_seen=checkpoint_seen + processed_since_start,
                last_item_id=last_processed_id,
                message=safe_json(result.errors[:3]) if result.errors else "crawl completed",
            )
            try:
                self.db.finish_batch(batch_id, status, safe_json(result.__dict__))
            except Exception:
                pass
            close_handler = getattr(handler, "close", None)
            if callable(close_handler):
                try:
                    close_handler()
                except Exception:
                    pass

    # 鈹鈹 澧為噺閲囬泦杈呭姪鏂规硶 鈹鈹
    def _query_existing_ids(self, platform: str) -> set[str]:
        q = getattr(self.db, "query_existing_source_item_ids", None)
        if callable(q):
            try:
                return set(q(platform))
            except Exception:
                return set()
        return set()

    def _query_list_fingerprints(self, platform: str) -> dict[str, str]:
        q = getattr(self.db, "query_list_fingerprints", None)
        if callable(q):
            try:
                return q(platform) or {}
            except Exception:
                return {}
        return {}

    def _update_list_fingerprints(self, platform: str, items: list[Any], handler: PlatformHandler) -> None:
        fp_method = getattr(handler, "list_fingerprint", None)
        rows: list[dict[str, Any]] = []
        for item in items:
            sid = _list_item_id(item)
            if not sid:
                continue
            fp = fp_method(item) if callable(fp_method) else None
            rows.append({
                "source_platform": platform,
                "source_item_id": sid,
                "fingerprint": fp or "",
                "updated_at": jd_v2.now_text(),
            })
        u = getattr(self.db, "upsert_list_fingerprints", None)
        if callable(u) and rows:
            try:
                u(rows)
            except Exception:
                pass

    def _crawl_list_item(
        self,
        handler: PlatformHandler,
        batch_id: str,
        list_item: Any,
        result: Optional[PlatformCrawlResult] = None,
    ) -> None:
        try:
            detail_bundle = handler.fetch_detail(list_item)
            record = handler.build_record(detail_bundle)
            if self.ai_enabled and self.ai_mode == "sync":
                self._apply_ai(record, handler, detail_bundle)
            # 鍐欏叆(MySQL 姝婚攣/閿佺瓑寰呮椂鑷姩閲嶈瘯, 淇濊瘉鍏ㄩ噺骞跺彂閲囬泦涓嶄涪鏁版嵁)
            self._write_record_with_retry(batch_id, record)
            if self.parse_attachments:
                self._parse_item_attachments(record)
            if self.ai_enabled and self.ai_mode == "async":
                self._enqueue_ai(record, handler, detail_bundle)
            if result is not None:
                result.success_count += 1
        except Exception as exc:
            err_msg = str(exc)
            # 閲囬泦澶辫触涔熷啓鍏ユ爣绾ч槦鍒? 渚涘叏閲忛噰闆嗚拷婧笌閲嶈窇
            try:
                self.db.write_crawl_queue_item(
                    batch_id=batch_id,
                    source_platform=getattr(handler, "source_platform", ""),
                    source_item_id=compact_text(getattr(list_item, "source_item_id", "") or list_item),
                    project_name=getattr(list_item, "title", None) or getattr(list_item, "project_name", None),
                    status="failed",
                    error_message=err_msg[:2000],
                )
            except Exception:
                pass
            if result is None:
                raise
            result.failed_count += 1
            result.errors.append(
                {
                    "item": compact_text(getattr(list_item, "source_item_id", "")) or compact_text(list_item),
                    "error": err_msg,
                }
            )

    def _apply_ai(self, record: PlatformRecord, handler: PlatformHandler, detail_bundle: Any) -> None:
        context_builder = getattr(getattr(handler, "adapter", None), "build_ai_context", None)
        if context_builder is None:
            return
        context = context_builder(detail_bundle)
        if not isinstance(context, AIExtractionContext):
            return
        context.asset_group = record.asset_group
        if not context.paimai_id:
            context.paimai_id = f"{record.source_platform}:{record.source_item_id}"
        ai_results = self._batch_extract_ai(record.asset_group, context)
        if not ai_results:
            return
        common_values, common_results, special_values, special_results = split_ai_results(ai_results, record.asset_group)
        for key, value in common_values.items():
            if not jd_v2.is_blank(value):
                if key in AI_FILL_ONLY_COMMON_FIELDS and not jd_v2.is_blank(record.common_values.get(key)):
                    continue
                record.common_values[key] = value
                record.field_results[key] = common_results[key]
        for key, value in special_values.items():
            if not jd_v2.is_blank(value):
                record.special_values[key] = value
                record.special_field_results[key] = special_results[key]
        self._apply_ai_detail_rows(record, ai_results)

    def _batch_extract_ai(self, asset_group: str, context: AIExtractionContext) -> Mapping[str, Any]:
        extractor = getattr(jd_v2, "ai_extractor", None)
        if extractor is None or not getattr(extractor, "is_available", lambda: False)():
            return {}
        return extractor.batch_extract(ai_fields_for_asset_group(asset_group), context)

    def _active_ai_runtime_info(self) -> dict[str, str]:
        extractor = getattr(jd_v2, "ai_extractor", None)
        if extractor is None:
            return {"profile_name": "", "provider": "", "model_name": ""}
        return {
            "profile_name": compact_text(
                getattr(extractor, "profile_name", "") or getattr(extractor, "ai_profile", "")
            ),
            "provider": compact_text(getattr(extractor, "provider", "")),
            "model_name": compact_text(getattr(extractor, "model_name", "")),
        }

    def _enqueue_ai(self, record: PlatformRecord, handler: PlatformHandler, detail_bundle: Any) -> None:
        context_builder = getattr(getattr(handler, "adapter", None), "build_ai_context", None)
        if context_builder is None or not hasattr(self.db, "enqueue_ai_enrichment_task"):
            return
        context = context_builder(detail_bundle)
        if not isinstance(context, AIExtractionContext):
            return
        context.asset_group = record.asset_group
        if not context.paimai_id:
            context.paimai_id = f"{record.source_platform}:{record.source_item_id}"
        self.db.enqueue_ai_enrichment_task(
            paimai_id=record.source_item_id,
            source_platform=record.source_platform,
            source_item_id=record.source_item_id,
            asset_group=record.asset_group,
            context=ai_context_to_payload(context),
            task_type="field_enrichment",
            priority=100,
            reason="main crawl wrote item first; queued AI enrichment",
        )

    def _parse_item_attachments(self, record: PlatformRecord) -> None:
        """Download and parse item attachments into text resources."""
        try:
            import hashlib
            import tempfile

            attachments = record.attachments_json
            if not attachments:
                return

            # 鎻愬彇鎵鏈夐檮浠?URL
            files: list[dict[str, Any]] = []
            if isinstance(attachments, dict):
                files = attachments.get("files") or attachments.get("data") or []
            elif isinstance(attachments, list):
                files = attachments

            if not files:
                return

            for attachment in files:
                if not isinstance(attachment, dict):
                    continue
                # 鑾峰彇鏂囦欢鍚嶅拰 URL
                name = compact_text(attachment.get("attachmentName") or attachment.get("name") or attachment.get("fileName") or "")
                url = compact_text(attachment.get("attachmentAddress") or attachment.get("url") or attachment.get("href") or attachment.get("downloadUrl") or "")
                if not url:
                    continue

                # 纭畾鏂囦欢绫诲瀷锛堣烦杩囦笉闇瑕佽В鏋愮殑绫诲瀷锛?
                ext = ""
                if name:
                    ext = os.path.splitext(name)[1].lower()
                type_name = attachment.get("attachmentType") or attachment.get("type") or ""

                # 鍙鐞嗘枃鏈被闄勪欢
                supported = ext in (".pdf", ".doc", ".docx", ".xls", ".xlsx", ".txt", ".csv", ".htm", ".html", ".rtf")
                if not supported and type_name not in ("pdf", "doc", "docx", "xls", "xlsx", "txt"):
                    continue

                # 涓嬭浇闄勪欢
                try:
                    content = self._download_raw(url)
                    if not content:
                        continue
                except Exception as download_err:
                    jd_v2.logger.warning("attachment_download_failed",
                        url=url[:120], name=name, error=str(download_err))
                    continue

                # 鎻愬彇鏂囨湰
                try:
                    text = self._extract_attachment_text(content, name)
                    if not text or len(text.strip()) < 10:
                        continue
                except Exception as extract_err:
                    jd_v2.logger.warning("attachment_text_extract_failed",
                        name=name, error=str(extract_err))
                    continue

                # 瀛樺偍涓?item_resource
                text_hash = hashlib.md5(text.encode("utf-8")).hexdigest()[:16]
                try:
                    raw_payload = {
                        "type": "attachment_text",
                        "filename": name,
                        "url": url,
                        "text_hash": text_hash,
                        "text_size": len(text),
                        "created_by": "parse_attachments",
                    }
                    if hasattr(self.db, "upsert_raw_payloads"):
                        self.db.upsert_raw_payloads(
                            source_platform=record.source_platform,
                            source_item_id=record.source_item_id,
                            payloads=[raw_payload],
                        )
                    else:
                        # Fallback: 浣跨敤 DB 鐨勯氱敤鏂规硶
                        from jd_mysql_store import MySQLJDScraperDatabase
                        if isinstance(self.db, MySQLJDScraperDatabase):
                            self.db.raw_insert("item_resources", {
                                "source_platform": record.source_platform,
                                "source_item_id": record.source_item_id,
                                "resource_type": "attachment_text",
                                "resource_name": name or url[:80],
                                "resource_url": url,
                                "resource_content": text[:65535],
                                "created_by": "parse_attachments",
                            })

                    jd_v2.logger.info("attachment_parsed",
                        platform=record.source_platform,
                        item=record.source_item_id,
                        name=name,
                        text_chars=len(text))
                except Exception as store_err:
                    jd_v2.logger.warning("attachment_store_failed",
                        name=name, error=str(store_err))

        except Exception as e:
            jd_v2.logger.warning("parse_item_attachments_error",
                item=record.source_item_id, error=str(e))

    @staticmethod
    def _download_raw(url: str) -> Optional[bytes]:
        """Download raw attachment bytes."""
        import requests as _requests
        try:
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            }
            resp = _requests.get(url, headers=headers, timeout=30, stream=True)
            if resp.status_code != 200:
                return None
            content = b""
            for chunk in resp.iter_content(chunk_size=65536):
                content += chunk
                if len(content) > 50 * 1024 * 1024:  # 50MB 涓婇檺
                    return None
            return content if content else None
        except Exception:
            return None

    @staticmethod
    def _extract_attachment_text(content: bytes, filename: str) -> Optional[str]:
        """Extract text from attachment bytes."""
        if not content:
            return None
        ext = os.path.splitext(filename or "unknown.txt")[1].lower()
        try:
            if ext == ".txt" or ext == ".csv" or ext == ".rtf":
                # 灏濊瘯澶氱缂栫爜
                for enc in ("utf-8", "gbk", "gb2312", "latin-1"):
                    try:
                        return content.decode(enc)
                    except UnicodeDecodeError:
                        continue
                return content.decode("utf-8", errors="replace")

            if ext == ".htm" or ext == ".html":
                text = content.decode("utf-8", errors="replace")
                # 绠鍗曞幓闄?HTML 鏍囩
                import re as _re
                text = _re.sub(r"<style[^>]*>.*?</style>", "", text, flags=_re.DOTALL | _re.IGNORECASE)
                text = _re.sub(r"<script[^>]*>.*?</script>", "", text, flags=_re.DOTALL | _re.IGNORECASE)
                text = _re.sub(r"<[^>]+>", " ", text)
                text = _re.sub(r"\s+", " ", text).strip()
                return text if len(text) > 10 else None

            if ext == ".pdf":
                try:
                    import PyPDF2
                    import io
                    reader = PyPDF2.PdfReader(io.BytesIO(content))
                    pages = []
                    for page in reader.pages:
                        page_text = page.extract_text()
                        if page_text:
                            pages.append(page_text)
                    return "\n\n".join(pages) if pages else None
                except ImportError:
                    pass  # pyPDF2 not available
                try:
                    import pdfplumber
                    import io
                    with pdfplumber.open(io.BytesIO(content)) as pdf:
                        pages = [page.extract_text() or "" for page in pdf.pages]
                    return "\n\n".join(pages) if pages else None
                except ImportError:
                    pass

            if ext in (".docx", ".doc"):
                try:
                    import docx
                    import io
                    doc = docx.Document(io.BytesIO(content))
                    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
                    return "\n".join(paragraphs)
                except ImportError:
                    pass

            if ext in (".xlsx", ".xls"):
                try:
                    import openpyxl
                    import io
                    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
                    rows_data = []
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        rows_data.append(f"--- Sheet: {sheet_name} ---")
                        for row in ws.iter_rows(values_only=True):
                            row_text = "\t".join(str(c) if c is not None else "" for c in row)
                            if row_text.strip():
                                rows_data.append(row_text)
                    wb.close()
                    return "\n".join(rows_data[:5000])
                except ImportError:
                    pass

            # 鍏朵粬绫诲瀷锛氬皾璇?UTF-8 瑙ｇ爜
            try:
                text = content.decode("utf-8")
                if len(text) > 10:
                    return text
            except UnicodeDecodeError:
                pass

            return None
        except Exception:
            return None

    def process_ai_enrichment_queue(
        self,
        *,
        limit: int = 20,
        worker_id: str = "ai-worker",
        concurrency: int = 1,
        task_types: Optional[list[str]] = None,
    ) -> dict[str, Any]:
        """Consume AI enrichment queue tasks."""
        if not hasattr(self.db, "fetch_ai_enrichment_tasks"):
            raise AttributeError("db does not support ai enrichment queue")
        normalized_task_types = _parse_task_types_arg(task_types)
        tasks = self.db.fetch_ai_enrichment_tasks(limit=limit, worker_id=worker_id, task_types=normalized_task_types)
        summary: dict[str, Any] = {
            "picked": len(tasks),
            "success": 0,
            "failed": 0,
            "skipped": 0,
            "errors": [],
            "task_types": normalized_task_types,
        }
        if not tasks:
            return summary
        summary_lock = threading.Lock()

        def process_one(task: dict[str, Any]) -> None:
            task_id = int(task["ai_task_id"])
            source_platform = compact_text(task.get("source_platform")) or "jd"
            source_item_id = compact_text(task.get("source_item_id"))
            asset_group = compact_text(task.get("asset_group")) or "other"
            try:
                context = ai_context_from_payload(task.get("context_json"))
                context.asset_group = asset_group
                if not context.paimai_id:
                    context.paimai_id = f"{source_platform}:{source_item_id}"
                if hasattr(self.db, "mark_ai_enrichment_task_parsing"):
                    runtime = self._active_ai_runtime_info()
                    can_parse = self.db.mark_ai_enrichment_task_parsing(
                        task_id,
                        worker_id=worker_id,
                        profile_name=runtime["profile_name"],
                        provider=runtime["provider"],
                        model_name=runtime["model_name"],
                    )
                    if not can_parse:
                        with summary_lock:
                            summary["skipped"] += 1
                        return
                ai_results = self._batch_extract_ai(asset_group, context)
                if not ai_results:
                    raise RuntimeError("AI extractor unavailable or returned no enrichment result")

                # 鈹鈹 Vision AI: 鐭ヨ瘑浜ф潈鏍囩殑鐨勫浘鐗囪〃鏍?OCR 鍏滃簳 鈹鈹鈹鈹鈹鈹鈹鈹鈹鈹
                # batch_extract 鍙仛鏂囨湰鎻愬彇锛屼笉浼犲浘鐗囩粰瑙嗚妯″瀷銆?
                # 褰撴爣鐨勭被鍨嬩负鐭ヨ瘑浜ф潈(ip) 涓斾笂涓嬫枃涓湁鍥剧墖 URL 鏃讹紝
                # 棰濆璋冪敤瑙嗚妯″瀷鎻愬彇鍥剧墖琛ㄦ牸涓殑閫愰」鏄庣粏銆?
                if asset_group == "ip" and context.image_urls:
                    extractor = getattr(jd_v2, "ai_extractor", None)
                    vision_fn = getattr(extractor, "extract_ip_details_from_images", None)
                    if vision_fn:
                        try:
                            image_result = vision_fn(context.image_urls, context)
                            if image_result and image_result.value:
                                ai_results["ip_details"] = image_result
                        except Exception:
                            logger.warning("async_vision_ip_failed",
                                           "寮傛瑙嗚 IP 鎻愬彇澶辫触",
                                           paimai_id=source_item_id)
                common_values, common_results, special_values, special_results = split_ai_results(ai_results, asset_group)
                record = PlatformRecord(
                    source_platform=source_platform,
                    source_site_name=source_platform,
                    source_item_id=source_item_id,
                    source_url="",
                    asset_group=asset_group,
                    common_values=common_values,
                    field_results=common_results,
                    special_values=special_values,
                    special_field_results=special_results,
                )
                self._apply_ai_detail_rows(record, ai_results)
                if not record.common_values and not record.special_values and not record.debt_details and not record.ip_details:
                    raise RuntimeError("AI extractor returned only empty/error enrichment values")
                self.db.apply_ai_enrichment_results(
                    paimai_id=source_item_id,
                    source_platform=source_platform,
                    asset_group=asset_group,
                    common_values=record.common_values,
                    common_results=record.field_results,
                    special_values=record.special_values,
                    special_results=record.special_field_results,
                    debt_details=record.debt_details,
                    ip_details=record.ip_details,
                )
                self.db.mark_ai_enrichment_task_success(task_id, ai_results_to_payload(ai_results))
                with summary_lock:
                    summary["success"] += 1
            except Exception as exc:
                with summary_lock:
                    summary["failed"] += 1
                    summary["errors"].append({"ai_task_id": task_id, "error": str(exc)})
                if hasattr(self.db, "mark_ai_enrichment_task_failed"):
                    self.db.mark_ai_enrichment_task_failed(task_id, exc)

        concurrency = max(1, int(concurrency or 1))
        if concurrency <= 1 or len(tasks) <= 1:
            for task in tasks:
                process_one(task)
        else:
            with ThreadPoolExecutor(max_workers=min(concurrency, len(tasks))) as executor:
                list(executor.map(process_one, tasks))
        return summary

    def _apply_ai_detail_rows(self, record: PlatformRecord, ai_results: Mapping[str, Any]) -> None:
        if record.asset_group == "debt":
            result = ai_results.get("debt_package_details_json")
            details = jd_v2.normalize_ai_debt_details(getattr(result, "value", None) if result else None)
            if not details:
                return
            if record.debt_details:
                details = record.debt_details
            else:
                record.debt_details = details
            derived_values = {
                "household_count": jd_v2.debt_detail_household_count(details),
                "benchmark_date": jd_v2.debt_detail_first_benchmark_date(details),
                "debtor_name": jd_v2.debt_detail_primary_debtor_names(details),
            }
            for key, value in derived_values.items():
                if jd_v2.is_blank(value) or not jd_v2.is_blank(record.special_values.get(key)):
                    continue
                record.special_values[key] = value
                record.special_field_results[key] = make_field_result(
                    value,
                    "ai_extraction",
                    f"llm_batch.{key}_from_debt_details",
                    getattr(result, "original_text", "") if result else value,
                    method="ai_derived",
                    confidence=getattr(result, "confidence", 0.75) if result else 0.75,
                )
        elif record.asset_group == "ip":
            result = ai_results.get("ip_details")
            details = jd_v2.normalize_ai_ip_details(getattr(result, "value", None) if result else None)
            if not details or jd_v2.ip_details_look_aggregated(details):
                details = self._single_ip_detail_from_special_values(record)
                if not details:
                    return
            record.ip_details = details
            if jd_v2.is_blank(record.special_values.get("ip_count")):
                value = str(len(details))
                record.special_values["ip_count"] = value
                record.special_field_results["ip_count"] = make_field_result(
                    value,
                    "ai_extraction",
                    "llm_batch.ip_count_from_ip_details",
                    getattr(result, "original_text", "") if result else value,
                    method="ai_derived",
                    confidence=getattr(result, "confidence", 0.75) if result else 0.75,
                )

    def _single_ip_detail_from_special_values(self, record: PlatformRecord) -> list[dict[str, Any]]:
        values = record.special_values or {}
        ip_name = compact_text(values.get("subject_name")) or ""
        certificate_no = compact_text(values.get("certificate_no")) or ""
        ip_type = compact_text(values.get("ip_type") or values.get("specific_category")) or ""
        if not any((ip_name, certificate_no, ip_type)):
            return []
        if re.search(r"\d+\s*(?:\u9879|\u4ef6)", ip_name) or re.search(r"\d+\s*(?:\u9879|\u4ef6)", certificate_no):
            return []
        excerpt_parts: list[str] = []
        for key in ("subject_name", "certificate_no", "ip_type", "specific_category"):
            result = record.special_field_results.get(key) or {}
            excerpt = compact_text(result.get("source_excerpt") or result.get("source_value")) or ""
            if excerpt and excerpt not in excerpt_parts:
                excerpt_parts.append(excerpt)
        return [
            {
                "sequence_no": "1",
                "ip_name": ip_name or None,
                "certificate_no": certificate_no or None,
                "ip_type": ip_type or None,
                "right_holder": compact_text(values.get("right_holder")) or None,
                "right_status": compact_text(values.get("right_status")) or None,
                "source_excerpt": "; ".join(excerpt_parts[:4]) or None,
            }
        ]

    def _write_record_with_retry(self, batch_id: str, record: PlatformRecord, max_retries: int = 5) -> None:
        """Write a record with retry for transient MySQL lock errors."""
        import time as _time
        last_exc: Optional[Exception] = None
        for attempt in range(1, max_retries + 1):
            try:
                self._write_record(batch_id, record)
                return
            except Exception as exc:
                msg = str(exc)
                if "1213" in msg or "1205" in msg or "Deadlock" in msg or "Lock wait timeout" in msg:
                    last_exc = exc
                    # 鎸囨暟閫閬?0.2s~2s)鍚庨噸璇? 闄嶄綆骞跺彂鍐欏叆鍐茬獊姒傜巼
                    _time.sleep(min(0.2 * (2 ** (attempt - 1)), 2.0))
                    continue
                raise
        if last_exc:
            raise last_exc

    def _write_record(self, batch_id: str, record: PlatformRecord) -> None:
        source_item_id = compact_text(record.source_item_id)
        if not source_item_id:
            raise ValueError("source_item_id is required")
        common_values = dict(record.common_values)
        common_values.update(
            {
                "source_platform": record.source_platform,
                "source_item_id": source_item_id,
                "source_url": record.source_url,
                "source_site_name": record.source_site_name,
            }
        )
        if "data_source" not in common_values or not compact_text(common_values.get("data_source")):
            common_values["data_source"] = record.source_site_name
        if "project_name" not in common_values or not compact_text(common_values.get("project_name")):
            common_values["project_name"] = source_item_id
        if "attachments_json" not in common_values and record.attachments_json is not None:
            common_values["attachments_json"] = safe_json(record.attachments_json)

        raw = record.raw_payloads or {}
        self.db.upsert_raw_payloads(
            paimai_id=source_item_id,
            batch_id=batch_id,
            source_url=record.source_url,
            source_platform=record.source_platform,
            source_item_id=source_item_id,
            source_site_name=record.source_site_name,
            list_json=raw.get("list_json") or raw.get("list_html") or {},
            detail_json=raw.get("detail_json") or {},
            product_basic_json=raw.get("product_basic_json") or raw.get("auxiliary_json") or {},
            realtime_json=raw.get("realtime_json") or raw.get("status_json") or {},
            description_html=raw.get("description_html") or raw.get("detail_html") or "",
            notice_html=raw.get("notice_html") or "",
            announcement_html=raw.get("announcement_html") or "",
            attachments_json=record.attachments_json,
            vendor_json=raw.get("vendor_json") or {},
        )
        self.db.upsert_common_item(
            paimai_id=source_item_id,
            batch_id=batch_id,
            asset_group=record.asset_group,
            jd_category_id=record.category_id,
            jd_category_name=record.category_name,
            values=common_values,
            field_results=record.field_results,
            special_values=record.special_values,
        )
        if record.asset_group in jd_v2.SPECIAL_FIELDS:
            self.db.upsert_special_item(
                paimai_id=source_item_id,
                source_platform=record.source_platform,
                asset_group=record.asset_group,
                values=record.special_values,
                field_results=record.special_field_results,
            )
        if record.debt_details:
            self.db.upsert_debt_details(
                paimai_id=source_item_id,
                source_platform=record.source_platform,
                details=record.debt_details,
            )
        if record.ip_details:
            self.db.upsert_ip_details(
                paimai_id=source_item_id,
                source_platform=record.source_platform,
                details=record.ip_details,
            )


class RequestsHTMLClient:
    def __init__(self, *, timeout: int | float | None = 0, headers: Mapping[str, str] | None = None) -> None:
        self.session = requests.Session()
        self.session.headers.update(dict(DEFAULT_HEADERS))
        if headers:
            self.session.headers.update(dict(headers))
        self.timeout = request_timeout_value(timeout)

    def get_text(self, url: str) -> str:
        response = self.session.get(url, timeout=self.timeout)
        response.raise_for_status()
        if not response.encoding or response.encoding.lower() in {"iso-8859-1", "ascii"}:
            response.encoding = response.apparent_encoding or "utf-8"
        return response.text

    def post_text(self, url: str, data: Mapping[str, Any] | None = None) -> str:
        response = self.session.post(url, data=data, timeout=self.timeout)
        response.raise_for_status()
        if not response.encoding or response.encoding.lower() in {"iso-8859-1", "ascii"}:
            response.encoding = response.apparent_encoding or "utf-8"
        return response.text


class Ejy365LiveHandler:
    def list_fingerprint(self, item: Ejy365ListItem) -> str:
        """Build a stable list fingerprint for incremental change checks."""
        def _safe(v: Any) -> str:
            if v is None:
                return ""
            t = compact_text(v)
            return t if t is not None else ""
        parts = [
            _list_item_id(item),
            item.title,
            item.price_raw,
            item.status,
            item.signup_deadline,
            item.region,
        ]
        return "|".join(_safe(p) for p in parts)

    source_platform = "ejy365"
    source_platform = "ejy365"
    source_site_name = "e浜ゆ槗"

    def __init__(self, *, request_timeout: int | float | None = 0, project_types: Iterable[str] | None = None) -> None:
        self.adapter = Ejy365Adapter()
        self.client = RequestsHTMLClient(timeout=request_timeout)
        self.project_types = tuple(project_types or DEFAULT_EJY365_PROJECT_TYPES)

    def fetch_list(self, limit: int) -> list[Ejy365ListItem]:
        """Fetch list items. limit=0 means full crawl."""
        items: list[Ejy365ListItem] = []
        seen: set[str] = set()
        import time as _time

        max_pages = 200 if not limit else 5
        per_type_limit = (max(1, limit) + len(self.project_types) - 1) // max(1, len(self.project_types)) if limit else None

        # 鎸夌被鍨嬮愪竴缈婚〉锛岄伩鍏嶅绉嶇被鍨嬩氦鍙夐犳垚澶ч噺璇锋眰
        for project_type in self.project_types:
            type_count = 0
            for page in range(1, max_pages + 1):
                if per_type_limit and type_count >= per_type_limit:
                    break
                url = f"{EJY365_BASE_URL}/jygg_more?project_type={project_type}&page={page}"
                try:
                    html = self.client.get_text(url)
                except Exception:
                    break  # 姝ょ被鍨嬪悗缁〉澶辫触锛岃烦鍒颁笅涓绉嶇被鍨?
                page_items = self.adapter.parse_list_html(html, base_url=EJY365_BASE_URL)
                if not page_items:
                    break  # 绌洪〉锛屾棤鏇村鏁版嵁
                added = 0
                for item in page_items:
                    key = item.slug or item.detail_url
                    if key in seen:
                        continue
                    seen.add(key)
                    setattr(item, "project_type_code", project_type)
                    items.append(item)
                    type_count += 1
                    added += 1
                    if limit and len(items) >= limit:
                        return items
                    if per_type_limit and type_count >= per_type_limit:
                        break
                if page > 1:
                    _time.sleep(0.3)  # 椤甸棿鐭欢杩熼槻鍙嶇埇
                if added == 0:
                    break
        return items

    def fetch_detail(self, list_item: Ejy365ListItem) -> Ejy365DetailBundle:
        html = self.client.get_text(list_item.detail_url)
        bundle = self.adapter.parse_detail_html(html, url=list_item.detail_url, list_item=list_item)
        # 灏濊瘯閫氳繃 jmjl_detail API 鑾峰彇绔炰拱璁板綍鍜岃ˉ鍏呬俊鎭?
        if bundle.raw_payloads.get("detail_html"):
            infoid = self.adapter._extract_infoid(bundle.raw_payloads["detail_html"])
            if infoid:
                jmjl_data = self.adapter.fetch_jmjl_detail(infoid)
                if jmjl_data:
                    bundle = self.adapter.parse_detail_html(
                        html, url=list_item.detail_url, list_item=list_item,
                        jmjl_detail=jmjl_data,
                    )
        return bundle

    def build_record(self, detail_bundle: Ejy365DetailBundle) -> PlatformRecord:
        common = self.adapter.map_common_candidates(detail_bundle)
        field_results = common.pop("field_results", {})
        source_item_id = compact_text(common.get("source_item_id") or detail_bundle.source_item_id or detail_bundle.url)
        attachments = normalize_attachments_payload(
            detail_bundle.attachments,
            [{"imageVideoArea": {"imageList": [{"imagePath": url} for url in detail_bundle.image_urls]}}],
        )
        project_type = compact_text(getattr(detail_bundle.list_item, "project_type_code", "")) if detail_bundle.list_item else "ZQ"
        asset_group, asset_label = ejy365_asset_for_project_type(project_type or "ZQ")
        common["source_site_name"] = self.source_site_name
        common["asset_group"] = asset_group
        common["asset_type"] = asset_label
        field_results["asset_group"] = make_field_result(
            asset_group,
            "list_html",
            "project_type",
            f"project_type={project_type or 'ZQ'}",
            method="category_mapping",
            confidence=0.9,
        )
        field_results["asset_type"] = make_field_result(
            asset_label,
            "list_html",
            "project_type",
            f"project_type={project_type or 'ZQ'}",
            method="category_mapping",
            confidence=0.9,
        )
        common["attachments_json"] = safe_json(attachments)
        if detail_bundle.bid_records_json is not None:
            common["bid_records_json"] = safe_json(detail_bundle.bid_records_json)
            field_results["bid_records_json"] = make_field_result(
                common["bid_records_json"],
                "detail_json",
                "jmjl_detail.bid_records",
                common["bid_records_json"],
                method="api",
                confidence=0.95,
            )
        if detail_bundle.jmjl_detail and detail_bundle.jmjl_detail.get("gg", {}).get("orgLXR"):
            contact_from_api = detail_bundle.jmjl_detail["gg"].get("orgLXR", "")
            phone_from_api = detail_bundle.jmjl_detail["gg"].get("orgPhone", "")
            if contact_from_api and not common.get("contact_info"):
                common["contact_info"] = f"{contact_from_api} {phone_from_api}".strip()
        special_values = self.adapter.map_special_candidates(detail_bundle, asset_group)
        special_field_results = special_results_from_values(asset_group, special_values, "detail_html", "ejy365")
        debt_details = self.adapter.extract_debt_details(detail_bundle) if asset_group == "debt" else []
        return PlatformRecord(
            source_platform=self.source_platform,
            source_site_name=self.source_site_name,
            source_item_id=source_item_id,
            source_url=detail_bundle.url,
            asset_group=asset_group,
            category_id=project_type or "ZQ",
            category_name=asset_label,
            common_values=common,
            field_results=field_results or common_results_from_values(common, "detail_html", "ejy365"),
            special_values=special_values,
            special_field_results=special_field_results,
            raw_payloads={
                "list_json": {"raw_html": detail_bundle.list_item.raw_html if detail_bundle.list_item else ""},
                "detail_html": detail_bundle.html,
                "auxiliary_json": detail_bundle.auxiliary_json or {},
                "status_json": detail_bundle.status_json or {},
                "jmjl_detail": detail_bundle.jmjl_detail or {},
            },
            attachments_json=attachments,
            debt_details=debt_details,
        )


class CquaeLiveHandler:
    def list_fingerprint(self, item: CquaeListItem) -> str:
        """Build a stable list fingerprint for incremental change checks."""
        def _safe(v: Any) -> str:
            if v is None:
                return ""
            t = compact_text(v)
            return t if t is not None else ""
        parts = [
            _list_item_id(item),
            item.title,
            item.project_status,
            item.price_raw,
            item.date_text,
            item.project_type,
        ]
        return "|".join(_safe(p) for p in parts)

    source_platform = CQUAE_PLATFORM
    source_platform = CQUAE_PLATFORM
    source_site_name = "???????"

    def __init__(
        self,
        *,
        request_timeout: int | float | None = 0,
        use_browser: bool = True,
        browser_headless: bool = True,
        browser_profile_path: Optional[str] = None,
        page_size: int = 60,
        max_pages: int = 0,
        browser_timeout_ms: int = 0,
        browser_settle_ms: int = 800,
    ) -> None:
        self.adapter = CquaeAdapter()
        self.client = RequestsHTMLClient(timeout=request_timeout)
        self.use_browser = use_browser
        self.page_size = max(1, int(page_size or 60))
        self.max_pages = max(0, int(max_pages or 0))
        self.browser = (
            CquaeBrowserFetcher(
                headless=browser_headless,
                timeout_ms=browser_timeout_ms,
                profile_path=browser_profile_path,
                settle_ms=browser_settle_ms,
            )
            if use_browser
            else None
        )
        # Inject pre-extracted WAF bypass cookies from .env (if available).
        # Key format: CQUAE_COOKIE__jsl_clearance_s  /  CQUAE_COOKIE_ASP.NET_SessionId
        import os
        for env_key, val in os.environ.items():
            if env_key.startswith("CQUAE_COOKIE_") and val:
                cookie_name = env_key[len("CQUAE_COOKIE_"):]
                self.client.session.cookies.set(cookie_name, val, domain=".cquae.com")

    def close(self) -> None:
        if self.browser:
            try:
                self.browser.close()
            except Exception:
                pass

    def _fetch_html(self, url: str) -> str:
        try:
            html = self.client.get_text(url)
            if not self.adapter.is_waf_challenge(html):
                return html
        except Exception:
            if not self.use_browser:
                raise
        if not self.browser:
            raise RuntimeError("CQUAE returned WAF challenge and browser fallback is disabled")
        rendered_html = self.browser.fetch_html(url)
        if self.adapter.is_waf_challenge(rendered_html):
            raise RuntimeError(
                "CQUAE returned a WAF challenge after browser fallback; "
                "try --cquae-headed with a trusted desktop browser session or switch to an official/API data source"
            )
        # 娴忚鍣ㄦ垚鍔熸覆鏌撳悗锛屾彁鍙栨柊椴滅殑 WAF cookies 娉ㄥ叆 HTTP session锛?
        # 鍚庣画璇锋眰鐩存帴鐢ㄦ柊 cookie 璧?HTTP锛屾棤闇鍐嶅紑娴忚鍣ㄣ?
        fresh = self.browser.last_waf_cookies
        if fresh:
            for name, value in fresh.items():
                self.client.session.cookies.set(name, value, domain=".cquae.com")
        return rendered_html

    def fetch_list(self, limit: int) -> list[CquaeListItem]:
        items: list[CquaeListItem] = []
        for page_items in self.iter_list_pages(limit):
            items.extend(page_items)
            if limit and len(items) >= limit:
                return items[:limit]
        if items:
            print(f"[CQUAE] fetched {len(items)} list items")
        return items

    def iter_list_pages(self, limit: int = 0):
        seen: set[str] = set()
        # Same asset may be listed under multiple project IDs on CQUAE.
        # Deduplicate by (title, price) to avoid collecting the same item
        # multiple times with different source_item_id values.
        first_error: Optional[Exception] = None
        max_pages = self.max_pages if self.max_pages > 0 else (1000 if not limit else 20)
        yielded = 0

        # 閲囬泦涓ょ被鐩細浜ф潈杞(projectID=1)銆佽祫浜ц浆璁?projectID=3)
        # 浜ф潈杞闇瑕嗙洊"姝ｅ紡鎶湶"+"棰勬姭闇?锛歯t=1/nt=8 鍧囦负姝ｅ紡鎶湶鎬?閮ㄥ垎閲嶅彔),
        # nt=3 涓洪鎶湶鎬?涓庢寮?0 閲嶅彔)锛涗笁鑰呴亶鍘嗗苟鎸?source_item_id 鍘婚噸銆?
        category_configs = [
            (1, 1, "equity_transfer_official"),
            (1, 8, "equity_transfer_official"),
            (1, 3, "equity_transfer_preview"),
            (3, None, "asset_transfer"),
        ]

        for project_id, nt_val, category_label in category_configs:
            nt = nt_val if nt_val is not None else 1
            try:
                # 鍏堝皾璇曚紶缁熺殑 URL 缈婚〉锛堝鏈嶅姟绔垎椤垫湁鏁堬級
                url = self.adapter.build_list_url(page=1, page_size=self.page_size, project_id=project_id, nt=nt, price_id=32)
                html = self._fetch_html(url)
                page_items = self.adapter.parse_list_html(html, base_url=CQUAE_BASE_URL)
                print(f"[CQUAE] {category_label} (projectID={project_id}, nt={nt_val}) page 1 returned {len(page_items)} items")
                if page_items:
                    fresh_items: list[CquaeListItem] = []
                    for item in page_items:
                        if item.source_item_id in seen:
                            continue
                        seen.add(item.source_item_id)
                        fresh_items.append(item)
                        yielded += 1
                        if limit and yielded >= limit:
                            break
                    if fresh_items:
                        yield fresh_items
                    if limit and yielded >= limit:
                        return
                    # 灏濊瘯 URL 缈婚〉锛坧age=2,3,鈥︼級
                    for page in range(2, max_pages + 1):
                        url = self.adapter.build_list_url(page=page, page_size=self.page_size, project_id=project_id, nt=nt, price_id=32)
                        try:
                            html2 = self._fetch_html(url)
                        except Exception:
                            break
                        page_items2 = self.adapter.parse_list_html(html2, base_url=CQUAE_BASE_URL)
                        if not page_items2:
                            break
                        print(f"[CQUAE] {category_label} (projectID={project_id}, nt={nt_val}) page {page} returned {len(page_items2)} items")
                        fresh_items = []
                        for item in page_items2:
                            if item.source_item_id in seen:
                                continue
                            seen.add(item.source_item_id)
                            fresh_items.append(item)
                            yielded += 1
                            if limit and yielded >= limit:
                                break
                        if not fresh_items:
                            break
                        yield fresh_items
                        if limit and yielded >= limit:
                            return
                else:
                    # URL 缈婚〉绗竴椤靛氨杩斿洖绌?鈫?鍙兘鏄?JS 鍒嗛〉锛?
                    # 鐢?Playwright 娓叉煋鍒楄〃椤碉紝鐐瑰嚮"涓嬩竴椤?缈婚〉
                    if self.browser:
                        browser_items = self._fetch_list_via_browser_click(
                            project_id, nt, category_label, [], seen, limit, max_pages
                        )
                        if browser_items:
                            yielded += len(browser_items)
                            yield browser_items
                        if limit and yielded >= limit:
                            return
            except Exception as exc:
                if first_error is None:
                    first_error = exc
                continue
        if first_error:
            raise first_error

    def _fetch_list_via_browser_click(
        self,
        project_id: int,
        nt: int,
        category_label: str,
        items: list[CquaeListItem],
        seen: set[str],
        limit: int,
        max_pages: int,
    ) -> list[CquaeListItem]:
        """Render list pages with Playwright and paginate."""
        from playwright.sync_api import sync_playwright
        url = self.adapter.build_list_url(page=1, page_size=self.page_size, project_id=project_id, nt=nt, price_id=32)
        headless = bool(getattr(self.browser, "headless", True))
        profile_path = getattr(self.browser, "profile_path", None)
        timeout_ms = int(getattr(self.browser, "timeout_ms", 30000) or 0)
        settle_ms = int(getattr(self.browser, "settle_ms", 800) or 0)
        with sync_playwright() as p:
            ctx = (
                p.chromium.launch_persistent_context(profile_path, headless=headless)
                if profile_path
                else p.chromium.launch(headless=headless, channel="chrome")
            )
            page = ctx.new_page()
            if timeout_ms <= 0:
                page.set_default_timeout(0)
                page.set_default_navigation_timeout(0)
            page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms or 0)
            for _ in range(max_pages):
                if settle_ms:
                    page.wait_for_timeout(settle_ms)
                html = page.content()
                page_items = self.adapter.parse_list_html(html, base_url=CQUAE_BASE_URL)
                if not page_items:
                    break
                added = 0
                for item in page_items:
                    if item.source_item_id in seen:
                        continue
                    seen.add(item.source_item_id)
                    items.append(item)
                    added += 1
                    if limit and len(items) >= limit:
                        ctx.close()
                        return items
                if added == 0:
                    break
                # 灏濊瘯鐐瑰嚮鈥滀笅涓椤碘?
                try:
                    next_btn = page.query_selector(".pagination .next, .pager .next, a.next, a:has-text('涓嬩竴椤?)")
                    if not next_btn:
                        next_btn = page.query_selector("a[rel=next]")
                    if not next_btn:
                        # 閫氱敤锛氭壘鏈鍚庝竴涓潪绂佺敤鍒嗛〉閾炬帴
                        links = page.query_selector_all(".pagination a, .pager a, .page a")
                        current = page.query_selector(".pagination .active, .pager .active, .page .active")
                        if links and current:
                            cur_idx = None
                            for i, lnk in enumerate(links):
                                if lnk.evaluate("el => el.classList.contains('active')"):
                                    cur_idx = i
                                    break
                            if cur_idx is not None and cur_idx < len(links) - 1:
                                next_btn = links[cur_idx + 1]
                    if next_btn and next_btn.is_enabled():
                        next_btn.click()
                        page.wait_for_load_state("domcontentloaded", timeout=timeout_ms or 10000)
                    else:
                        break
                except Exception:
                    break
            ctx.close()
        return items

    def fetch_detail(self, list_item: CquaeListItem) -> CquaeDetailBundle:
        html = self._fetch_html(list_item.source_url)
        return self.adapter.parse_detail_html(html, url=list_item.source_url, list_item=list_item)

    def build_record(self, detail_bundle: CquaeDetailBundle) -> PlatformRecord:
        common = self.adapter.map_common_candidates(detail_bundle)
        adapter_field_results = common.pop("field_results", {})
        asset_group = common.get("asset_group") or self.adapter.classify_bundle(detail_bundle) or "other"
        special = self.adapter.map_special_candidates(detail_bundle, asset_group)
        attachments = normalize_attachments_payload(
            detail_bundle.attachments,
            [{"imageVideoArea": {"imageList": [{"imagePath": url} for url in detail_bundle.image_urls]}}],
        )
        common["attachments_json"] = safe_json(attachments)
        field_results = common_results_from_values(common, "detail_html", "cquae")
        for key, fr in adapter_field_results.items():
            field_results[key] = fr
        special_results = special_results_from_values(asset_group, special, "detail_html", "cquae")
        source_item_id = compact_text(detail_bundle.source_item_id)
        record_source_site_name = compact_text(common.get("source_site_name")) or self.source_site_name
        return PlatformRecord(
            source_platform=self.source_platform,
            source_site_name=record_source_site_name,
            source_item_id=source_item_id,
            source_url=detail_bundle.source_url,
            asset_group=asset_group,
            category_id="cquae",
            category_name=compact_text(common.get("asset_type")) or "浜ф潈浜ゆ槗",
            common_values=common,
            field_results=field_results,
            special_values=special,
            special_field_results=special_results,
            raw_payloads={
                "list_json": detail_bundle.list_item.raw_fields if detail_bundle.list_item else {},
                "detail_html": detail_bundle.raw_html,
            },
            attachments_json=attachments,
        )


class SdcqjyLiveHandler:
    def list_fingerprint(self, item: CquaeListItem) -> str:
        """Build a stable list fingerprint for incremental change checks."""
        def _safe(v: Any) -> str:
            if v is None:
                return ""
            t = compact_text(v)
            return t if t is not None else ""
        parts = [
            _list_item_id(item),
            item.title,
            item.project_status,
            item.price_raw,
            item.date_text,
            item.project_type,
        ]
        return "|".join(_safe(p) for p in parts)

    source_platform = SDCQJY_PLATFORM
    source_platform = SDCQJY_PLATFORM
    source_site_name = SDCQJY_DATA_SOURCE

    def __init__(self, *, request_timeout: int | float | None = 0) -> None:
        self.adapter = SdcqjyAdapter()
        self.client = RequestsHTMLClient(timeout=request_timeout)

    def fetch_list(self, limit: int) -> list[CquaeListItem]:
        items: list[CquaeListItem] = []
        seen: set[str] = set()

        # 鍏堣闂椤佃幏鍙?session cookie
        try:
            self.client.get_text("http://www.sdcqjy.com/")
        except Exception:
            pass

        # 浠呴噰闆嗕笁涓垎绫伙細浜ф潈(cq)銆佽祫浜?zc)銆佽瘔璁肩綒娌?ssfm)
        type_ids = ["cq", "zc", "ssfm"]
        max_pages = 100 if not limit else 5
        page_size = 15  # API 姣忛〉鍥哄畾 15 鏉?

        for type_id in type_ids:
            for page in range(1, max_pages + 1):
                try:
                    html = self.client.post_text(SDCQJY_LIST_ENDPOINT, data={
                        "categoryId": "xmpd",
                        "typeId": type_id,
                        "page": page,
                        "projType": "table",
                    })
                except Exception:
                    break
                page_items = self.adapter.parse_list_html(html)
                if not page_items:
                    break
                added = 0
                for item in page_items:
                    if item.source_item_id in seen:
                        continue
                    seen.add(item.source_item_id)
                    items.append(item)
                    added += 1
                    if limit and len(items) >= limit:
                        return items
                if added == 0:
                    break
        return items

    def fetch_detail(self, list_item: CquaeListItem) -> CquaeDetailBundle:
        html = self.client.get_text(list_item.source_url)
        return self.adapter.parse_detail_html(html, url=list_item.source_url, list_item=list_item)

    def build_record(self, detail_bundle: CquaeDetailBundle) -> PlatformRecord:
        common = self.adapter.map_common_candidates(detail_bundle)
        adapter_field_results = common.pop("field_results", {})
        common["source_platform"] = self.source_platform
        common["source_site_name"] = self.source_site_name
        common["data_source"] = self.source_site_name
        asset_group = common.get("asset_group") or self.adapter.classify_bundle(detail_bundle) or "other"
        special = self.adapter.map_special_candidates(detail_bundle, asset_group)
        attachments = normalize_attachments_payload(
            detail_bundle.attachments,
            [{"imageVideoArea": {"imageList": [{"imagePath": url} for url in detail_bundle.image_urls]}}],
        )
        common["attachments_json"] = safe_json(attachments)
        field_results = common_results_from_values(common, "detail_html", self.source_platform)
        for key, fr in adapter_field_results.items():
            field_results[key] = fr
        special_results = special_results_from_values(asset_group, special, "detail_html", self.source_platform)
        source_item_id = compact_text(detail_bundle.source_item_id)
        return PlatformRecord(
            source_platform=self.source_platform,
            source_site_name=self.source_site_name,
            source_item_id=source_item_id,
            source_url=detail_bundle.source_url,
            asset_group=asset_group,
            category_id=self.source_platform,
            category_name=compact_text(common.get("asset_type")) or "灞变笢浜ф潈浜ゆ槗",
            common_values=common,
            field_results=field_results,
            special_values=special,
            special_field_results=special_results,
            raw_payloads={
                "list_json": detail_bundle.list_item.raw_fields if detail_bundle.list_item else {},
                "detail_html": detail_bundle.raw_html,
            },
            attachments_json=attachments,
        )


class AliLiveHandler:
    def list_fingerprint(self, item: AliListItem) -> str:
        """Build a stable list fingerprint for incremental change checks."""
        def _safe(v: Any) -> str:
            if v is None:
                return ""
            t = compact_text(v)
            return t if t is not None else ""
        parts = [
            _list_item_id(item),
            item.title,
            item.project_status,
            item.start_price_raw,
            item.final_price_raw,
        ]
        return "|".join(_safe(p) for p in parts)

    source_platform = ALI_SOURCE_PLATFORM
    source_platform = ALI_SOURCE_PLATFORM
    source_site_name = "闃块噷鎷嶅崠"

    def __init__(
        self,
        *,
        profile_path: Optional[str] = None,
        item_urls: Iterable[str] | None = None,
        headless: bool = False,
        timeout_ms: int = 0,
        tk_token: Optional[str] = None,
    ) -> None:
        self.adapter = AliAuctionAdapter()
        self.profile_path = profile_path
        self.item_urls = list(item_urls or [])
        self.headless = headless
        self.timeout_ms = timeout_ms
        # If a pre-extracted _m_h5_tk token is provided (via CLI or .env),
        # inject it directly into the MTOP session so no browser is needed.
        if not tk_token:
            import os
            tk_token = os.environ.get("ALI_TK_TOKEN") or ""
        if tk_token:
            self.adapter.mtop_fetcher.session.cookies.set(
                "_m_h5_tk", tk_token, domain=".taobao.com",
            )

    def fetch_list(self, limit: int) -> list[AliListItem]:
        if self.item_urls:
            return [self._list_item_from_url(url) for url in (self.item_urls[:limit] if limit else self.item_urls)]

        # 鈹鈹 Path 1: Bootstrap MTOP session cookies via Playwright 鈹鈹鈹鈹鈹鈹鈹鈹鈹鈹
        # The taobao auction homepage is a DataFront SPA 鈥?it does NOT contain
        # <a> tags pointing to individual auction items.  The only way to get a
        # listing is through the MTOP API, which requires the _m_h5_tk cookie
        # for signing.  We use Playwright + Chrome profile purely to obtain
        # that cookie from the logged-in browser session.
        if self.profile_path:
            try:
                from playwright.sync_api import sync_playwright
            except ImportError:
                pass
            else:
                try:
                    with sync_playwright() as playwright:
                        context = playwright.chromium.launch_persistent_context(
                            user_data_dir=self.profile_path,
                            headless=self.headless,
                            channel="chrome",
                        )
                        page = context.new_page()
                        page.goto(
                            "https://zc-paimai.taobao.com/",
                            wait_until="domcontentloaded",
                            timeout=self.timeout_ms or 30000,
                        )
                        import time
                        time.sleep(3)
                        # Transfer cookies from the browser session into the
                        # MTOP requests session so API signing succeeds.
                        browser_cookies = context.cookies()
                        found_tk = False
                        for cookie in browser_cookies:
                            if cookie["name"] in ("_m_h5_tk", "_m_h5_tk_enc"):
                                self.adapter.mtop_fetcher.session.cookies.set(
                                    cookie["name"],
                                    cookie["value"],
                                    domain=cookie.get("domain", ".taobao.com"),
                                )
                                found_tk = True
                        logger.info("ali_fetch_list playwright got cookies",
                                    f"count={len(browser_cookies)}, _m_h5_tk={found_tk}")
                        context.close()
                    items = self.adapter.fetch_mtop_list(
                        limit=limit,
                        channels=ALI_LIST_CHANNELS,
                        pages_per_channel=100 if not limit else 3,
                    )
                    logger.info("ali_fetch_list mtop returned items",
                                str(len(items)))
                    if items:
                        return items
                except Exception:
                    import traceback
                    logger.error("ali_fetch_list path1 failed",
                                 traceback.format_exc())

        # 鈹鈹 Path 2: MTOP API (may also work if warm-up got the cookie) 鈹鈹鈹鈹
        try:
            items = self.adapter.fetch_mtop_list(
                limit=limit,
                channels=ALI_LIST_CHANNELS,
                pages_per_channel=100 if not limit else 3,
            )
            if items:
                return items
        except Exception:
            pass

        # 鈹鈹 Path 3: Browser fallback (Playwright/Selenium) 鈹鈹鈹鈹鈹鈹鈹鈹鈹鈹鈹鈹鈹鈹鈹鈹
        return self._fetch_list_with_browser(limit)

    def _list_item_from_url(self, url: str) -> AliListItem:
        item_id = self._extract_item_id(url)
        return AliListItem(item_id=item_id, source_url=url, title=item_id, raw={"url": url})

    def _extract_item_id(self, url: str) -> str:
        for pattern in (
            r"(?:[?&](?:(?:itemId|item_id)|auctionId)=)(\d+)",
            r"(?:[?&]id=)(\d+)",
        ):
            match = re.search(pattern, url)
            if match:
                return match.group(1)
        digits = re.findall(r"\d{6,}", url)
        return digits[-1] if digits else url

    def _fetch_list_with_browser(self, limit: int) -> list[AliListItem]:
        try:
            from playwright.sync_api import sync_playwright
        except ImportError:
            return self._fetch_list_with_selenium(limit)
        with sync_playwright() as playwright:
            context = playwright.chromium.launch_persistent_context(
                user_data_dir=self.profile_path,
                headless=self.headless,
                channel="chrome",
            )
            try:
                page = context.new_page()
                page.goto("https://zc-paimai.taobao.com/", wait_until="domcontentloaded", timeout=self.timeout_ms)
                import time
                time.sleep(5)
                hrefs = page.eval_on_selector_all("a[href]", "els => els.map(a => a.href)")
            finally:
                context.close()
        return self._list_items_from_page(hrefs, "", limit)

    def _fetch_list_with_selenium(self, limit: int) -> list[AliListItem]:
        try:
            from selenium import webdriver
            from selenium.common.exceptions import TimeoutException, WebDriverException
        except ImportError as exc:
            raise RuntimeError("Ali live listing requires Playwright or Selenium.") from exc

        last_error: Optional[Exception] = None
        for browser_name, driver_factory, options_factory in (
            ("chrome", webdriver.Chrome, webdriver.ChromeOptions),
            ("edge", webdriver.Edge, webdriver.EdgeOptions),
        ):
            options = options_factory()
            options.page_load_strategy = "eager"
            if self.headless:
                options.add_argument("--headless=new")
            options.add_argument("--disable-gpu")
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            options.add_argument("--window-size=1365,900")
            options.add_argument(
                "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"
            )
            if self.profile_path and browser_name == "chrome":
                options.add_argument(f"--user-data-dir={self.profile_path}")
            driver = None
            try:
                driver = driver_factory(options=options)
                if self.timeout_ms and self.timeout_ms > 0:
                    driver.set_page_load_timeout(max(1, self.timeout_ms / 1000))
                try:
                    driver.get("https://zc-paimai.taobao.com/")
                except TimeoutException:
                    driver.execute_script("window.stop()")
                time.sleep(5)
                hrefs = driver.execute_script(
                    "return Array.from(document.querySelectorAll('a[href]')).map(a => a.href)"
                )
                return self._list_items_from_page(list(hrefs or []), driver.page_source or "", limit)
            except WebDriverException as exc:
                last_error = exc
            finally:
                if driver is not None:
                    try:
                        driver.quit()
                    except Exception:
                        pass
        raise RuntimeError(f"Ali live listing failed with browser fallback: {last_error}")

    def _list_items_from_page(self, hrefs: Iterable[str], html: str, limit: int) -> list[AliListItem]:
        candidates: list[str] = []
        candidates.extend(str(href) for href in hrefs if href)
        candidates.extend(re.findall(r"https?://[^\"'\\s<>]+auction\\.htm[^\"'\\s<>]+", html or "", flags=re.I))
        for match in re.findall(r"(?:(?:itemId|item_id))[\"'=:\\s]+(\\d{5,})", html or "", flags=re.I):
            candidates.append(ALI_DEFAULT_DETAIL_URL.format(item_id=match))

        items: list[AliListItem] = []
        seen: set[str] = set()
        for href in candidates:
            if "itemId=" not in href and "auction.htm" not in href:
                continue
            item_id = self._extract_item_id(href)
            if not item_id or item_id in seen:
                continue
            seen.add(item_id)
            source_url = href if href.startswith("http") else ALI_DEFAULT_DETAIL_URL.format(item_id=item_id)
            items.append(AliListItem(item_id=item_id, source_url=source_url, title=item_id, raw={"url": source_url}))
            if len(items) >= limit:
                break
        if not items:
            raise RuntimeError(
                "Ali listing rendered but no auction item links were found; "
                "the page likely requires login/session cookies or an mtop list API."
            )
        return items

    def fetch_detail(self, list_item: AliListItem) -> AliDetailBundle:
        url = list_item.source_url or ALI_DEFAULT_DETAIL_URL.format(item_id=list_item.item_id)
        try:
            bundle = self.adapter.fetch_mtop_detail(list_item)
            try:
                browser_bundle = self.adapter.browser_fetcher.fetch_detail(
                    url,
                    profile_path=self.profile_path,
                    timeout_ms=self.timeout_ms,
                )
                bundle = self.adapter.merge_detail_bundles(bundle, browser_bundle)
            except Exception:
                pass
        except Exception:
            bundle = self.adapter.browser_fetcher.fetch_detail(
                url,
                profile_path=self.profile_path,
                timeout_ms=self.timeout_ms,
            )
        bundle.list_item = list_item
        if not bundle.source_item_id:
            bundle.source_item_id = list_item.item_id
        if not bundle.source_url:
            bundle.source_url = url
        return bundle

    def build_record(self, detail_bundle: AliDetailBundle) -> PlatformRecord:
        if detail_bundle.status != "ok":
            raise RuntimeError(f"Ali detail blocked: {detail_bundle.status} {detail_bundle.block_reason}")
        common = self.adapter.map_common_candidates(detail_bundle)
        adapter_field_results = common.pop("field_results", {})
        asset_group = detail_bundle.asset_group or common.get("asset_group") or "other"
        special = self.adapter.map_special_candidates(detail_bundle, asset_group)
        common["source_site_name"] = self.source_site_name
        attachments = normalize_attachments_payload(detail_bundle.attachments, [{"imageVideoArea": {"imageList": [{"imagePath": url} for url in detail_bundle.image_urls]}}])
        common["attachments_json"] = safe_json(attachments)
        field_results = common_results_from_values(common, "detail_html", "ali")
        for key, fr in adapter_field_results.items():
            field_results[key] = fr
        special_results = special_results_from_values(asset_group, special, "detail_html", "ali")
        source_item_id = compact_text(detail_bundle.source_item_id)
        return PlatformRecord(
            source_platform=self.source_platform,
            source_site_name=self.source_site_name,
            source_item_id=source_item_id,
            source_url=detail_bundle.source_url,
            asset_group=asset_group,
            category_id="ali",
            category_name=detail_bundle.category or compact_text(common.get("asset_type")),
            common_values=common,
            field_results=field_results,
            special_values=special,
            special_field_results=special_results,
            raw_payloads={
                "list_json": detail_bundle.list_item.raw if detail_bundle.list_item else {},
                "detail_html": detail_bundle.rendered_html,
                "notice_html": detail_bundle.notice_html,
                "announcement_html": detail_bundle.notice_html,
                "detail_json": detail_bundle.top_json or {},
            },
            attachments_json=attachments,
        )


class TpreLiveHandler:
    def list_fingerprint(self, item: TpreListItem) -> str:
        """Build a stable list fingerprint for incremental change checks."""
        def _safe(v: Any) -> str:
            if v is None:
                return ""
            t = compact_text(v)
            return t if t is not None else ""
        parts = [
            _list_item_id(item),
            item.title,
            item.price_raw,
            item.project_status_name,
            item.end_time,
            item.biz_type_name,
        ]
        return "|".join(_safe(p) for p in parts)

    source_platform = TPRE_PLATFORM
    source_platform = TPRE_PLATFORM
    source_site_name = TPRE_DATA_SOURCE

    def __init__(self, *, request_timeout: int | float | None = 0) -> None:
        self.adapter = TpreAdapter(timeout=request_timeout if request_timeout else 15)

    def fetch_list(self, limit: int) -> list[TpreListItem]:
        items: list[TpreListItem] = []
        seen: set[str] = set()

        # 澶氬垎绫婚亶鍘嗭細浠呴噰姝ｅ紡鎶湶锛團ORMAL锛屾湁浜ゆ槗浠锋牸锛?
        system_codes = list(TpreAdapter.TPRE_SYSTEM_CODES.keys())
        for system_code in system_codes:
            max_pages = 200 if not limit else 5
            page_size = min(limit, 20) if limit else 50
            for page in range(1, max_pages + 1):
                api_data = self.adapter.fetch_list_api(
                    page=page, size=page_size,
                    system_code=system_code, biz_type_code="FORMAL",
                )
                page_items = self.adapter.parse_list_response(api_data)
                if not page_items:
                    break
                added = 0
                for item in page_items:
                    if item.source_item_id in seen:
                        continue
                    seen.add(item.source_item_id)
                    items.append(item)
                    added += 1
                    if limit and len(items) >= limit:
                        return items
                if added == 0:
                    break
        return items

    def fetch_detail(self, list_item: TpreListItem) -> TpreDetailBundle:
        api_data = self.adapter.fetch_detail_api(list_item)
        return self.adapter.parse_detail_response(api_data, list_item=list_item)

    def build_record(self, detail_bundle: TpreDetailBundle) -> PlatformRecord:
        common = self.adapter.map_common_candidates(detail_bundle)
        field_results = common.pop("field_results", {})
        asset_group_common = common.get("asset_group") or self.adapter.classify_bundle(detail_bundle) or "other"
        special = self.adapter.map_special_candidates(detail_bundle, asset_group_common)
        common["source_site_name"] = self.source_site_name
        attachments = normalize_attachments_payload(
            detail_bundle.attachments,
            [{"imageVideoArea": {"imageList": [{"imagePath": url} for url in detail_bundle.image_urls]}}],
        )
        common["attachments_json"] = safe_json(attachments)
        common_results = common_results_from_values(common, "detail_api", TPRE_PLATFORM)
        special_results = special_results_from_values(asset_group_common, special, "detail_api", TPRE_PLATFORM)
        for key, fr in field_results.items():
            common_results.setdefault(key, fr)
        asset_type = compact_text(common.get("asset_type")) or "浜ф潈杞"
        source_item_id = compact_text(detail_bundle.source_item_id)
        return PlatformRecord(
            source_platform=self.source_platform,
            source_site_name=self.source_site_name,
            source_item_id=source_item_id,
            source_url=detail_bundle.source_url,
            asset_group=asset_group_common,
            category_id=TPRE_PLATFORM,
            category_name=asset_type,
            common_values=common,
            field_results=common_results,
            special_values=special,
            special_field_results=special_results,
            raw_payloads={
                "list_json": asdict(detail_bundle.list_item) if detail_bundle.list_item else {},
                "detail_json": detail_bundle.detail_json,
                "detail_html": detail_bundle.raw_html,
            },
            attachments_json=attachments,
        )


class PrechinaLiveHandler:
    source_platform = PRECHINA_PLATFORM
    source_site_name = PRECHINA_DATA_SOURCE

    def __init__(self, *, request_timeout: int | float | None = 0) -> None:
        self.adapter = PrechinaAdapter(timeout=request_timeout if request_timeout else 20)
        self.client = RequestsHTMLClient(timeout=request_timeout)

    def fetch_list(self, limit: int) -> list[PrechinaListItem]:
        html = self.client.get_text(PRECHINA_BASE_URL)
        items = self.adapter.parse_list_from_homepage(html)
        return items[:limit] if limit else items

    def list_fingerprint(self, item: PrechinaListItem) -> str:
        """Build a stable list fingerprint for incremental change checks."""
        def _safe(v: Any) -> str:
            if v is None:
                return ""
            t = compact_text(v)
            return t if t is not None else ""
        parts = [
            item.source_item_id,
            item.title,
            item.price_raw,
            item.announce_date,
            item.end_date,
            item.status_code,
            item.biz_type_code,
        ]
        return "|".join(_safe(p) for p in parts)

    def fetch_detail(self, list_item: PrechinaListItem) -> PrechinaDetailBundle:
        if list_item.source_url and list_item.source_url != f"{PRECHINA_BASE_URL}/ejygg/index.jhtml":
            html = self.client.get_text(list_item.source_url)
        else:
            html = ""
        return self.adapter.parse_detail_html(html, url=list_item.source_url, list_item=list_item)

    def build_record(self, detail_bundle: PrechinaDetailBundle) -> PlatformRecord:
        common = self.adapter.map_common_candidates(detail_bundle)
        field_results = common.pop("field_results", {})
        asset_group_common = common.get("asset_group") or self.adapter.classify_bundle(detail_bundle) or "other"
        special = self.adapter.map_special_candidates(detail_bundle, asset_group_common)
        common["source_site_name"] = self.source_site_name
        attachments = normalize_attachments_payload(
            detail_bundle.attachments,
            [{"imageVideoArea": {"imageList": [{"imagePath": url} for url in detail_bundle.image_urls]}}],
        )
        common["attachments_json"] = safe_json(attachments)
        common_results = common_results_from_values(common, "detail_html", PRECHINA_PLATFORM)
        special_results = special_results_from_values(asset_group_common, special, "detail_html", PRECHINA_PLATFORM)
        for key, fr in field_results.items():
            common_results.setdefault(key, fr)
        asset_type = compact_text(common.get("asset_type")) or "浜ф潈杞"
        source_item_id = compact_text(detail_bundle.source_item_id)
        return PlatformRecord(
            source_platform=self.source_platform,
            source_site_name=self.source_site_name,
            source_item_id=source_item_id,
            source_url=detail_bundle.source_url,
            asset_group=asset_group_common,
            category_id=PRECHINA_PLATFORM,
            category_name=asset_type,
            common_values=common,
            field_results=common_results,
            special_values=special,
            special_field_results=special_results,
            raw_payloads={
                "list_json": detail_bundle.list_item.raw_fields if detail_bundle.list_item else {},
                "detail_html": detail_bundle.raw_html,
            },
            attachments_json=attachments,
        )


class GxcqLiveHandler:
    def list_fingerprint(self, item: GxcqListItem) -> str:
        """Build a stable list fingerprint for incremental change checks."""
        def _safe(v: Any) -> str:
            if v is None:
                return ""
            t = compact_text(v)
            return t if t is not None else ""
        parts = [
            _list_item_id(item),
            item.title,
            item.price_raw,
            item.project_status,
            item.end_time,
            item.assets_type_parent,
        ]
        return "|".join(_safe(p) for p in parts)

    source_platform = GXCQ_PLATFORM
    source_site_name = GXCQ_DATA_SOURCE

    def __init__(self, *, request_timeout: int | float | None = 0) -> None:
        self.adapter = GxcqAdapter(timeout=request_timeout if request_timeout else 15)
        self.client = RequestsHTMLClient(timeout=request_timeout)

    def fetch_list(self, limit: int) -> list[GxcqListItem]:
        """Fetch list items. limit=0 means full crawl."""
        items: list[GxcqListItem] = []
        seen: set[str] = set()
        page_size = 10 if not limit else min(max(1, limit), 10)

        max_pages = 1000  # 瀹夊叏涓婇檺锛岄槻姝㈡寰幆

        if not limit:
            # 鍏ㄩ噺妯″紡: 缈诲畬鎵鏈夊垎椤?
            # 鍏堣姹傜涓椤碉紝鑾峰彇鎬绘暟
            api_data = self.adapter.fetch_list_api(page=1, size=page_size)
            page_items = self.adapter.parse_list_response(api_data)
            total_count = self.adapter.parse_total_count(api_data)
            for item in page_items:
                if item.source_item_id in seen:
                    continue
                seen.add(item.source_item_id)
                items.append(item)

            if total_count > 0:
                # 宸茬煡鎬绘暟锛岀簿纭炕椤?
                total_pages = (total_count + page_size - 1) // page_size
                print(f"[GXCQ] total_count={total_count}, page_size={page_size}, total_pages={total_pages}")
                for page in range(2, total_pages + 1):
                    if len(items) >= max_pages * page_size:
                        break
                    # 閲嶈瘯3娆★紝闂撮殧閫掑
                    for attempt in range(3):
                        try:
                            api_data = self.adapter.fetch_list_api(page=page, size=page_size)
                            page_items = self.adapter.parse_list_response(api_data)
                            print(f"[GXCQ] page={page}, items_in_page={len(page_items)}, total_so_far={len(items)}")
                            if not page_items:
                                break
                            for item in page_items:
                                if item.source_item_id in seen:
                                    continue
                                seen.add(item.source_item_id)
                                items.append(item)
                            break  # success, exit retry loop
                        except Exception as e:
                            if attempt < 2:
                                import time
                                wait = (attempt + 1) * 3
                                print(f"[GXCQ] page={page} failed (attempt {attempt+1}/3): {e}, retry in {wait}s")
                                time.sleep(wait)
                            else:
                                print(f"[GXCQ] page={page} failed after 3 attempts: {e}")
                                # 璺宠繃杩欎竴椤电户缁笅涓椤?
                                break
            else:
                # 鏃犳硶鑾峰彇鎬绘暟锛岄愰〉缈荤洿鍒扮┖椤?
                print(f"[GXCQ] No total_count, falling back to max_pages={max_pages}")
                for page in range(2, max_pages + 1):
                    try:
                        api_data = self.adapter.fetch_list_api(page=page, size=page_size)
                        page_items = self.adapter.parse_list_response(api_data)
                        if not page_items:
                            break
                        added = 0
                        for item in page_items:
                            if item.source_item_id in seen:
                                continue
                            seen.add(item.source_item_id)
                            items.append(item)
                            added += 1
                        if added == 0:
                            break
                    except Exception:
                        break
            return items

        # 閲囨牱妯″紡: PHPCMF API
        for page in range(1, min(6, max_pages + 1)):
            api_data = self.adapter.fetch_list_api(page=page, size=page_size)
            page_items = self.adapter.parse_list_response(api_data)
            if not page_items:
                break
            added = 0
            for item in page_items:
                if item.source_item_id in seen:
                    continue
                seen.add(item.source_item_id)
                items.append(item)
                added += 1
                if len(items) >= limit:
                    return items
            if added == 0:
                break
        return items

    def fetch_detail(self, list_item: GxcqListItem) -> GxcqDetailBundle:
        detail_data = self.adapter.fetch_detail_api(list_item)
        return self.adapter.parse_detail_response(detail_data, list_item=list_item)

    def build_record(self, detail_bundle: GxcqDetailBundle) -> PlatformRecord:
        common = self.adapter.map_common_candidates(detail_bundle)
        field_results = common.pop("field_results", {})
        asset_group_common = common.get("asset_group") or self.adapter.classify_bundle(detail_bundle) or "other"
        special = self.adapter.map_special_candidates(detail_bundle, asset_group_common)
        common["source_site_name"] = self.source_site_name
        attachments = normalize_attachments_payload(
            detail_bundle.attachments,
            [{"imageVideoArea": {"imageList": [{"imagePath": url} for url in detail_bundle.image_urls]}}],
        )
        common["attachments_json"] = safe_json(attachments)
        common_results = common_results_from_values(common, "detail_api", GXCQ_PLATFORM)
        special_results = special_results_from_values(asset_group_common, special, "detail_api", GXCQ_PLATFORM)
        for key, fr in field_results.items():
            common_results.setdefault(key, fr)
        source_item_id = compact_text(detail_bundle.source_item_id)
        return PlatformRecord(
            source_platform=self.source_platform,
            source_site_name=self.source_site_name,
            source_item_id=source_item_id,
            source_url=detail_bundle.source_url,
            asset_group=asset_group_common,
            category_id=GXCQ_PLATFORM,
            category_name=compact_text(common.get("asset_type")) or "浜ф潈杞",
            common_values=common,
            field_results=common_results,
            special_values=special,
            special_field_results=special_results,
            raw_payloads={
                "list_json": detail_bundle.list_item.raw_json if detail_bundle.list_item else {},
                "detail_html": detail_bundle.raw_html,
            },
            attachments_json=attachments,
        )


class GycqLiveHandler:
    """Guizhou property exchange live handler."""




    source_platform = GYCQ_PLATFORM
    source_site_name = GYCQ_DATA_SOURCE

    def __init__(self, *, request_timeout: int | float | None = 0) -> None:
        self.adapter = GycqAdapter(timeout=request_timeout if request_timeout else 20)
        self.client = RequestsHTMLClient(timeout=request_timeout)

    def fetch_list(self, limit: int) -> list[GxcqListItem]:
        """Fetch list items. limit=0 means full crawl."""



        items: list[GxcqListItem] = []
        seen: set[str] = set()
        page = 1
        while True:
            api_data = self.adapter.fetch_list_api(page=page, size=30)
            page_items = self.adapter.parse_list_response(api_data)
            if not page_items:
                break
            added = 0
            for item in page_items:
                if item.source_item_id in seen:
                    continue
                seen.add(item.source_item_id)
                items.append(item)
                added += 1
                if limit and len(items) >= limit:
                    return items
            # 鏈〉鍒ゅ畾: 鏈〉杩斿洖鏉℃暟 < 30 鎴栧凡鏃犳柊澧?
            payload = api_data.get("data") or {}
            if added == 0 or len(page_items) < 30:
                break
            page += 1
            if page > 200:
                break
        return items

    def list_fingerprint(self, item: GxcqListItem) -> str:
        """Build a stable list fingerprint for incremental change checks."""
        def _safe(v: Any) -> str:
            if v is None:
                return ""
            t = compact_text(v)
            return t if t is not None else ""
        rj = item.raw_json or {}
        parts = [
            item.source_item_id,
            item.title,
            item.price_raw,
            item.project_status,
            rj.get("_announcementStart"),
            rj.get("_announcementEnd"),
            item.assets_type_parent,
        ]
        return "|".join(_safe(p) for p in parts)

    def fetch_detail(self, list_item: GxcqListItem) -> GxcqDetailBundle:
        detail_data = self.adapter.fetch_detail_api(list_item)
        return self.adapter.parse_detail_response(detail_data, list_item=list_item)

    def build_record(self, detail_bundle: GxcqDetailBundle) -> PlatformRecord:
        common = self.adapter.map_common_candidates(detail_bundle)
        field_results = common.pop("field_results", {})
        asset_group_common = common.get("asset_group") or self.adapter.classify_bundle(detail_bundle) or "other"
        special = self.adapter.map_special_candidates(detail_bundle, asset_group_common)
        common["source_site_name"] = self.source_site_name
        attachments = normalize_attachments_payload(
            detail_bundle.attachments,
            [{"imageVideoArea": {"imageList": [{"imagePath": url} for url in detail_bundle.image_urls]}}],
        )
        common["attachments_json"] = safe_json(attachments)
        common_results = common_results_from_values(common, "detail_api", GYCQ_PLATFORM)
        special_results = special_results_from_values(asset_group_common, special, "detail_api", GYCQ_PLATFORM)
        for key, fr in field_results.items():
            common_results.setdefault(key, fr)
        source_item_id = compact_text(detail_bundle.source_item_id)
        return PlatformRecord(
            source_platform=self.source_platform,
            source_site_name=self.source_site_name,
            source_item_id=source_item_id,
            source_url=detail_bundle.source_url,
            asset_group=asset_group_common,
            category_id=GYCQ_PLATFORM,
            category_name=compact_text(common.get("asset_type")) or "浜ф潈杞",
            common_values=common,
            field_results=common_results,
            special_values=special,
            special_field_results=special_results,
            raw_payloads={
                "list_json": detail_bundle.list_item.raw_json if detail_bundle.list_item else {},
                "detail_html": detail_bundle.raw_html,
            },
            attachments_json=attachments,
        )


class CbexLiveHandler:
    def list_fingerprint(self, item: CbexListItem) -> str:
        """Build a stable list fingerprint for incremental change checks."""
        def _safe(v: Any) -> str:
            if v is None:
                return ""
            t = compact_text(v)
            return t if t is not None else ""
        parts = [
            _list_item_id(item),
            item.title,
            item.price_raw,
            item.status,
            item.trade_type,
            item.region,
        ]
        return "|".join(_safe(p) for p in parts)

    source_platform = CBEX_PLATFORM
    source_platform = CBEX_PLATFORM
    source_site_name = CBEX_DATA_SOURCE

    def __init__(self, *, request_timeout: int | float | None = 0, browser_profile_path: str | None = None) -> None:
        self.adapter = CbexAdapter()
        self.browser = CbexBrowserFetcher(headless=True, timeout_ms=60000, profile_path=browser_profile_path)

    # ===== 鍒楄〃/璇︽儏璧版祻瑙堝櫒涓婁笅鏂?(缁曡繃 WAF, 鍗曠嚎绋嬫墽琛? =====
    def _fetch_category(self, biz: str, seen: set, items: list, limit: int) -> int:
        """Fetch one CBEX businessType category and return totalRecordNum."""
        page = 1
        total = 0
        max_pages = 2000
        while page <= max_pages:
            data = self.browser.api_search(business_type=biz, disclosure_type="", from_page=page, page_size=15)
            if not data:
                break
            inner = data.get("data") if isinstance(data.get("data"), dict) else data
            total = (inner or {}).get("totalRecordNum") or 0
            rows = (inner or {}).get("data") or []
            if not rows:
                break
            for it in rows:
                li = CbexAdapter.from_api_item(it)
                if not li:
                    continue
                if li.prj_id in seen:
                    continue
                seen.add(li.prj_id)
                items.append(li)
                if limit and len(items) >= limit:
                    return total
            if page * 15 >= total:
                break
            page += 1
            time.sleep(0.1)
        return total

    def fetch_list(self, limit: int) -> list[CbexListItem]:
        """Fetch CBEX list items from JSON API categories."""







        items: list[CbexListItem] = []
        seen: set[str] = set()
        # 鍏ㄩ噺 businessType 浠ｇ爜 (浠庣綉绔?HTML 鎶犲嚭, 鍚悗缁彂鐜扮殑 CAR 绛?
        # 鏈熬杩藉姞 "" 鍏滃簳: 鎺ュ彛鍦ㄤ笉鎸囧畾 businessType 鏃朵細杩斿洖棰濆鏁版嵁
        # (濡?椤圭洰鎺ㄤ粙"绾?4700+ 鏉?, 杩欎簺鏁版嵁涓嶅綊鍏ヤ笂杩颁换涓鏍囧噯绫荤洰
        categories = ["ZQ", "ZL", "GZ", "JC", "ZS", "SSZC", "SW", "TJ", "CAR", ""]
        grand_total = 0
        for biz in categories:
            t = self._fetch_category(biz, seen, items, limit)
            if t:
                grand_total += t
                print(f"[CBEX] businessType={biz} api returned {t} records")
            if limit and len(items) >= limit:
                break
        print(f"[CBEX] 鍒楄〃鎺ュ彛绱鎶撳彇 {len(items)} 鏉?(鎺ュ彛 totalRecordNum 鍚堣 {grand_total})")
        return items

    def fetch_detail(self, list_item: CbexListItem) -> CbexDetailBundle:
        """Fetch detail page HTML for one item URL."""
        html = self.browser.fetch_detail_html_by_url(list_item.detail_url)
        if not html:
            return self.adapter.parse_detail_html("", list_item.prj_id, list_item=list_item)
        bundle = self.adapter.parse_detail_html(html, list_item.prj_id, list_item=list_item)
        bundle.url = list_item.detail_url
        return bundle

    def build_record(self, detail_bundle: CbexDetailBundle) -> PlatformRecord:
        common = self.adapter.map_common_candidates(detail_bundle)
        field_results = common.pop("field_results", {})
        asset_group_common = common.get("asset_group") or self.adapter.classify_bundle(detail_bundle) or "other"
        special = self.adapter.map_special_candidates(detail_bundle, asset_group_common)
        common["source_site_name"] = self.source_site_name
        attachments = normalize_attachments_payload(
            detail_bundle.attachments,
            [{"imageVideoArea": {"imageList": [{"imagePath": url} for url in detail_bundle.image_urls]}}],
        )
        common["attachments_json"] = safe_json(attachments)
        common_results = common_results_from_values(common, "detail_html", CBEX_PLATFORM)
        special_results = special_results_from_values(asset_group_common, special, "detail_html", CBEX_PLATFORM)
        for key, fr in field_results.items():
            common_results.setdefault(key, fr)
        source_item_id = compact_text(detail_bundle.source_item_id)
        return PlatformRecord(
            source_platform=self.source_platform,
            source_site_name=self.source_site_name,
            source_item_id=source_item_id,
            source_url=detail_bundle.url,
            asset_group=asset_group_common,
            category_id=CBEX_PLATFORM,
            category_name=compact_text(common.get("asset_type")) or "浜ф潈杞",
            common_values=common,
            field_results=common_results,
            special_values=special,
            special_field_results=special_results,
            raw_payloads={
                "detail_html": detail_bundle.html,
            },
            attachments_json=attachments,
        )


class JdLiveHandler:
    source_platform = JD_SOURCE_PLATFORM
    source_site_name = JD_DATA_SOURCE

    def __init__(
        self,
        *,
        request_timeout: Optional[float] = None,
        output_dir: str | Path = Path("outputs") / "multi_platform_jd",
        categories: set[str] | None = None,
    ) -> None:
        timeout = None if request_timeout is None or request_timeout <= 0 else int(request_timeout)
        self.adapter = JDPlatformAdapter(timeout=timeout)
        self.output_dir = Path(output_dir)
        self.categories = categories

    def crawl_with_db(
        self,
        db: Any,
        limit: int,
        mode: str = "sample",
        ai_mode: str = "async",
        checkpoint_callback: Callable[..., None] | None = None,
        resume_checkpoint: Mapping[str, Any] | None = None,
    ) -> PlatformCrawlResult:
        self.output_dir.mkdir(parents=True, exist_ok=True)
        # 鍏ㄩ噺妯″紡: 鍘绘帀鎬绘潯鏁颁笂闄? 姣忕被涓婇檺鏀惧ぇ(浜笢褰撳墠浠呭彇 page=1, 澶氶〉缈婚〉灞炲悗缁寮?
        full_mode = mode == "full" or limit == 0
        summary = self.adapter.crawl_sample(
            db=db,
            per_category_limit=limit if not full_mode else 100000,
            output_dir=self.output_dir,
            categories=self.categories,
            total_limit=None if full_mode else limit,
            mode=mode,
            ai_mode=ai_mode,
            checkpoint_callback=checkpoint_callback,
            resume_checkpoint=resume_checkpoint,
        )
        errors = summary.get("errors") or []
        scanned = int(summary.get("items_seen") or 0)
        return PlatformCrawlResult(
            platform=self.source_platform,
            batch_id=compact_text(summary.get("batch_id")) or "",
            scanned_count=scanned,
            success_count=max(0, scanned - len(errors)),
            failed_count=len(errors),
            errors=[{"item": compact_text(err.get("paimai_id")), "error": compact_text(err.get("error"))} for err in errors],
        )


def build_handlers(args: argparse.Namespace) -> dict[str, PlatformHandler]:
    ali_urls: list[str] = []
    for url in getattr(args, "ali_item_url", None) or []:
        ali_urls.extend([part.strip() for part in str(url).split(",") if part.strip()])
    jd_categories = {part.strip() for part in (getattr(args, "jd_categories", "") or "").split(",") if part.strip()} or None
    ejy365_types = tuple(part.strip() for part in (getattr(args, "ejy365_types", "") or "").split(",") if part.strip()) or None
    return {
        "jd": JdLiveHandler(
            request_timeout=getattr(args, "request_timeout", 30),
            output_dir=getattr(args, "output_dir", "outputs"),
            categories=jd_categories,
        ),
        "ejy365": Ejy365LiveHandler(
            request_timeout=getattr(args, "request_timeout", 30),
            project_types=ejy365_types,
        ),
        "cquae": CquaeLiveHandler(
            request_timeout=getattr(args, "request_timeout", 30),
            use_browser=not getattr(args, "no_browser", False),
            browser_headless=not getattr(args, "cquae_headed", False),
            browser_profile_path=getattr(args, "cquae_profile_path", None) or None,
            page_size=getattr(args, "cquae_page_size", 60),
            max_pages=getattr(args, "cquae_max_pages", 0),
            browser_timeout_ms=getattr(args, "browser_timeout_ms", 0),
            browser_settle_ms=getattr(args, "cquae_browser_settle_ms", 800),
        ),
        "sdcqjy": SdcqjyLiveHandler(request_timeout=getattr(args, "request_timeout", 30)),
        "ali": AliLiveHandler(
            profile_path=getattr(args, "ali_profile_path", None),
            item_urls=ali_urls,
            headless=getattr(args, "ali_headless", True),
            timeout_ms=getattr(args, "browser_timeout_ms", 30000),
            tk_token=getattr(args, "ali_tk_token", None),
        ),
        "tpre": TpreLiveHandler(request_timeout=getattr(args, "request_timeout", 30)),
        "prechina": PrechinaLiveHandler(request_timeout=getattr(args, "request_timeout", 30)),
        "gxcq": GxcqLiveHandler(request_timeout=getattr(args, "request_timeout", 30)),
        "gycq": GycqLiveHandler(request_timeout=getattr(args, "request_timeout", 30)),
        "cbex": CbexLiveHandler(
            request_timeout=getattr(args, "request_timeout", 30),
            browser_profile_path=getattr(args, "cquae_profile_path", None) or None,
        ),
    }


def _md_table(headers: list[str], rows: list[list[Any]]) -> str:
    if not rows:
        return "_鏃犳暟鎹甠\n"
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(compact_text(value).replace("|", "\\|") for value in row) + " |")
    return "\n".join(lines) + "\n"


def generate_model_quality_report(
    config: MySQLConfig,
    *,
    output_dir: Path,
    run_results: list[dict[str, Any]] | None = None,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    report_path = output_dir / f"model_data_quality_report_{time.strftime('%Y%m%d_%H%M%S')}.md"
    with mysql_connection(config) as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) AS c FROM auction_items")
            total_items = int((cur.fetchone() or {}).get("c") or 0)
            cur.execute(
                """
                SELECT
                  source_platform,
                  COALESCE(source_site_name, source_platform) AS source_site_name,
                  COUNT(*) AS item_count,
                  SUM(CASE WHEN final_price_display IS NULL OR final_price_display='' THEN 1 ELSE 0 END) AS missing_final_price,
                  SUM(CASE WHEN start_price_display IS NULL OR start_price_display='' THEN 1 ELSE 0 END) AS missing_start_price,
                  SUM(CASE WHEN contact_info IS NULL OR contact_info='' THEN 1 ELSE 0 END) AS missing_contact,
                  SUM(CASE WHEN assessment_price_display IS NULL OR assessment_price_display='' THEN 1 ELSE 0 END) AS missing_assessment
                FROM auction_items
                GROUP BY source_platform, source_site_name
                ORDER BY source_platform
                """
            )
            platform_rows = cur.fetchall() or []
            cur.execute(
                """
                SELECT resource_type, COUNT(*) AS count_rows
                FROM item_resources
                GROUP BY resource_type
                ORDER BY resource_type
                """
            )
            resource_rows = cur.fetchall() or []
            cur.execute(
                """
                SELECT COUNT(*) AS c
                FROM item_resources
                WHERE resource_url IS NULL OR resource_url=''
                """
            )
            empty_resource_urls = int((cur.fetchone() or {}).get("c") or 0)
            cur.execute(
                """
                SELECT ai.source_platform, COUNT(*) AS item_count
                FROM auction_items ai
                LEFT JOIN item_resources ir
                  ON ir.item_id=ai.item_id AND ir.resource_type='attachment'
                WHERE ir.resource_id IS NULL
                GROUP BY ai.source_platform
                ORDER BY ai.source_platform
                """
            )
            no_attachment_rows = cur.fetchall() or []
            cur.execute(
                """
                SELECT field_namespace, field_key, field_label, COUNT(*) AS count_rows
                FROM field_extractions
                WHERE status <> 'extracted'
                GROUP BY field_namespace, field_key, field_label
                ORDER BY count_rows DESC, field_namespace, field_key
                LIMIT 20
                """
            )
            missing_field_rows = cur.fetchall() or []
            cur.execute(
                """
                SELECT queue_status AS status, COUNT(*) AS count_rows
                FROM ai_enrichment_queue
                GROUP BY queue_status
                ORDER BY queue_status
                """
            )
            ai_queue_rows = cur.fetchall() or []

    result_rows = []
    for item in run_results or []:
        result_rows.append(
            [
                item.get("platform"),
                item.get("scanned_count"),
                item.get("success_count"),
                item.get("failed_count"),
                len(item.get("errors") or []),
            ]
        )

    platform_table = _md_table(
        ["platform", "site", "item_count", "missing_final_price", "missing_start_price", "missing_contact", "missing_assessment"],
        [
            [
                row.get("source_platform"),
                row.get("source_site_name"),
                row.get("item_count"),
                row.get("missing_final_price"),
                row.get("missing_start_price"),
                row.get("missing_contact"),
                row.get("missing_assessment"),
            ]
            for row in platform_rows
        ],
    )
    resource_table = _md_table(
        ["璧勬簮绫诲瀷", "鏁伴噺"],
        [[row.get("resource_type"), row.get("count_rows")] for row in resource_rows],
    )
    no_attachment_table = _md_table(
        ["platform", "items_without_attachment_files"],
        [[row.get("source_platform"), row.get("item_count")] for row in no_attachment_rows],
    )
    missing_field_table = _md_table(
        ["namespace", "field_key", "field_label", "missing_or_abnormal_count"],
        [
            [row.get("field_namespace"), row.get("field_key"), row.get("field_label"), row.get("count_rows")]
            for row in missing_field_rows
        ],
    )
    ai_queue_table = _md_table(
        ["status", "count"],
        [[row.get("status"), row.get("count_rows")] for row in ai_queue_rows],
    )
    result_table = _md_table(
        ["platform", "scanned", "success", "failed", "error_count"],
        result_rows,
    )

    platform_table = _md_table(
        ["platform", "site", "item_count", "missing_final_price", "missing_start_price", "missing_contact", "missing_assessment"],
        [
            [
                row.get("source_platform"),
                row.get("source_site_name"),
                row.get("item_count"),
                row.get("missing_final_price"),
                row.get("missing_start_price"),
                row.get("missing_contact"),
                row.get("missing_assessment"),
            ]
            for row in platform_rows
        ],
    )
    resource_table = _md_table(
        ["璧勬簮绫诲瀷", "鏁伴噺"],
        [[row.get("resource_type"), row.get("count_rows")] for row in resource_rows],
    )
    no_attachment_table = _md_table(
        ["platform", "items_without_attachment_files"],
        [[row.get("source_platform"), row.get("item_count")] for row in no_attachment_rows],
    )
    missing_field_table = _md_table(
        ["namespace", "field_key", "field_label", "missing_or_abnormal_count"],
        [
            [row.get("field_namespace"), row.get("field_key"), row.get("field_label"), row.get("count_rows")]
            for row in missing_field_rows
        ],
    )
    ai_queue_table = _md_table(
        ["status", "count"],
        [[row.get("status"), row.get("count_rows")] for row in ai_queue_rows],
    )
    result_table = _md_table(
        ["platform", "scanned", "success", "failed", "error_count"],
        result_rows,
    )
    generated_at = time.strftime("%Y-%m-%d %H:%M:%S")
    report = f"""# Model Crawl Data Quality Report

Generated at: {generated_at}

## Current Run Results

{result_table}

## Summary

- Total items in database: {total_items}
- item_resources empty URL count: {empty_resource_urls}

## Missing Fields By Platform

{platform_table}

## Resource Storage

{resource_table}

## Items Without Attachment Files

This counts items with no `item_resources.resource_type='attachment'` rows. Images/videos alone are not counted as attachment files.

{no_attachment_table}

## Missing/Abnormal Fields Top 20

{missing_field_table}

## AI Queue

{ai_queue_table}
"""
    report_path.write_text(report, encoding="utf-8")
    return report_path


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Multi-platform auction crawler runner")
    sub = parser.add_subparsers(dest="command", required=True)
    crawl = sub.add_parser("crawl", help="crawl one or more non-JD platforms into MySQL")
    crawl.add_argument("--platform", choices=["jd", "ejy365", "cquae", "sdcqjy", "ali", "tpre", "prechina", "gxcq", "gycq", "cbex", "all"], default="all")
    crawl.add_argument("--limit", type=int, default=10)
    crawl.add_argument("--mode", choices=["sample", "full", "incremental"], default="sample",
                       help="sample: 閲囨牱鍓?N 鏉?榛樿); full: 缈诲畬鎵鏈夊垎椤靛仛鍏ㄩ噺閲囬泦; "
                            "incremental: 浠呴噰闆嗘柊澧?鍒楄〃鍙樻洿鐨勬爣鐨?鍩轰簬宸插叆搴揑D涓庡垪琛ㄦ寚绾?")
    crawl.add_argument("--output-dir", type=Path, default=Path("outputs") / "multi_platform", help="export/output directory")
    crawl.add_argument("--reset-db", action="store_true", help="drop and recreate MySQL V2 tables before crawling")
    crawl.add_argument("--confirm-reset-db", action="store_true", help="required with --reset-db to confirm destructive table drops")
    crawl.add_argument("--mysql-host", default="127.0.0.1")
    crawl.add_argument("--mysql-port", type=int, default=3306)
    crawl.add_argument("--mysql-user", default=os.getenv("MYSQL_USER", "root"))
    crawl.add_argument("--mysql-password", default=os.getenv("MYSQL_PASSWORD", "root"))
    crawl.add_argument("--mysql-database", default=os.getenv("MYSQL_DATABASE", "auction_data"))
    crawl.add_argument("--platform-concurrency", type=int, default=1, help="number of platforms to crawl concurrently")
    crawl.add_argument("--item-concurrency", type=int, default=1, help="number of items to process concurrently per platform")
    crawl.add_argument("--request-timeout", type=float, default=0, help="HTTP timeout seconds; 0 means no local timeout")
    crawl.add_argument("--browser-timeout-ms", type=int, default=0, help="browser navigation timeout; 0 means no timeout")
    crawl.add_argument("--no-browser", action="store_true", help="disable browser fallback for WAF pages")
    crawl.add_argument("--cquae-headed", action="store_true", help="run CQUAE browser fallback in visible headed mode")
    crawl.add_argument("--cquae-profile-path", default="", help="browser user-data-dir for CQUAE WAF fallback")
    crawl.add_argument("--cquae-page-size", type=int, default=60, help="CQUAE list page size; larger values reduce list requests when accepted")
    crawl.add_argument("--cquae-max-pages", type=int, default=0, help="max CQUAE list pages per category; 0 uses mode default")
    crawl.add_argument("--cquae-browser-settle-ms", type=int, default=800, help="CQUAE browser fallback wait after DOM load")
    crawl.add_argument("--no-ai", action="store_true", help="disable AI enrichment")
    crawl.add_argument("--parse-attachments", action="store_true", help="download and extract text from attachments during crawl")
    crawl.add_argument(
        "--ai-mode",
        choices=["sync", "async", "off"],
        default="async",
        help="AI extraction mode: sync blocks crawl; async queues enrichment after DB write; off disables AI",
    )
    crawl.add_argument("--ai-model", default="")
    crawl.add_argument("--ai-profile", default="")
    crawl.add_argument("--ai-provider", default="")
    crawl.add_argument("--ai-model-name", default="")
    crawl.add_argument("--ai-api-key", default="")
    crawl.add_argument("--ai-base-url", default="")
    crawl.add_argument("--ali-profile-path", default="")
    crawl.add_argument("--ali-item-url", action="append", help="Ali detail URL; repeatable or comma-separated")
    crawl.add_argument("--ali-headless", action="store_true")
    crawl.add_argument("--ali-tk-token", default="",
        help="_m_h5_tk cookie value for MTOP API signing (portable: no Chrome profile needed)")
    crawl.add_argument("--jd-categories", default="", help="JD category ids, comma-separated; only used when --platform jd")
    crawl.add_argument("--ejy365-types", default="", help="E浜ゆ槗 project types, comma-separated, e.g. ZQ,GQ,TD; empty=all 36 types")

    enrich = sub.add_parser("ai-enrich", help="process queued AI enrichment tasks")
    enrich.add_argument("--limit", type=int, default=20)
    enrich.add_argument("--worker-id", default="ai-worker")
    enrich.add_argument("--concurrency", type=int, default=1, help="number of queue tasks to process concurrently (AI calls are I/O bound)")
    enrich.add_argument("--output-dir", type=Path, default=Path("outputs") / "multi_platform", help="export/output directory")
    enrich.add_argument("--mysql-host", default="127.0.0.1")
    enrich.add_argument("--mysql-port", type=int, default=3306)
    enrich.add_argument("--mysql-user", default=os.getenv("MYSQL_USER", "root"))
    enrich.add_argument("--mysql-password", default=os.getenv("MYSQL_PASSWORD", "root"))
    enrich.add_argument("--mysql-database", default=os.getenv("MYSQL_DATABASE", "auction_data"))
    enrich.add_argument("--ai-model", default="")
    enrich.add_argument("--ai-profile", default="")
    enrich.add_argument("--ai-provider", default="")
    enrich.add_argument("--ai-model-name", default="")
    enrich.add_argument("--ai-api-key", default="")
    enrich.add_argument("--ai-base-url", default="")
    enrich.add_argument("--task-types", default="", help="AI queue task types for this worker, comma-separated")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    config = MySQLConfig(
        host=args.mysql_host,
        port=args.mysql_port,
        user=args.mysql_user,
        password=args.mysql_password,
        database=args.mysql_database,
    )
    db = MySQLJDScraperDatabase(config)
    if args.command == "ai-enrich":
        db.init_schema()
        jd_v2.init_ai_extractor(
            args.ai_provider or args.ai_model,
            args.ai_api_key,
            args.ai_base_url,
            model_name=args.ai_model_name,
            profile=args.ai_profile,
            mysql_config=config,
        )
        runner = MultiPlatformRunner(db=db, handlers={}, ai_enabled=True, ai_mode="sync")
        result = runner.process_ai_enrichment_queue(
            limit=args.limit,
            worker_id=args.worker_id,
            concurrency=args.concurrency,
            task_types=_parse_task_types_arg(args.task_types),
        )
        report_path = generate_model_quality_report(config, output_dir=args.output_dir, run_results=[])
        result["quality_report"] = str(report_path)
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
        return 0 if result.get("failed", 0) == 0 else 1

    if args.reset_db and not args.confirm_reset_db:
        print("--reset-db is destructive; add --confirm-reset-db to confirm table reset.")
        return 2
    if args.reset_db:
        from jd_mysql_store import require_db_reset_allowed
        require_db_reset_allowed()
        reset_mysql_tables(config)
    db.init_schema()
    db.seed_field_catalog()
    ai_mode = "off" if args.no_ai else args.ai_mode
    if ai_mode == "sync":
        jd_v2.init_ai_extractor(
            args.ai_provider or args.ai_model,
            args.ai_api_key,
            args.ai_base_url,
            model_name=args.ai_model_name,
            profile=args.ai_profile,
            mysql_config=config,
        )
    handlers = build_handlers(args)
    platforms = ["jd", "ejy365", "cquae", "sdcqjy", "ali", "tpre", "prechina", "gxcq", "gycq", "cbex"] if args.platform == "all" else [args.platform]
    runner = MultiPlatformRunner(
        db=db,
        handlers=handlers,
        ai_enabled=ai_mode != "off",
        ai_mode=ai_mode,
        item_concurrency=args.item_concurrency,
        parse_attachments=getattr(args, "parse_attachments", False),
    )
    if args.platform_concurrency > 1 and len(platforms) > 1:
        results = []
        with ThreadPoolExecutor(max_workers=min(args.platform_concurrency, len(platforms))) as executor:
            futures = {
                executor.submit(runner.crawl_platform, platform, args.limit, args.mode): platform
                for platform in platforms
            }
            for future in as_completed(futures):
                platform = futures[future]
                try:
                    results.append(future.result().__dict__)
                except Exception as exc:
                    results.append(
                        {
                            "platform": platform,
                            "batch_id": "",
                            "scanned_count": 0,
                            "success_count": 0,
                            "failed_count": args.limit,
                            "errors": [{"item": "platform", "error": str(exc)}],
                        }
                    )
    else:
        results = [runner.crawl_platform(platform, limit=args.limit, mode=args.mode).__dict__ for platform in platforms]
    report_path = generate_model_quality_report(config, output_dir=args.output_dir, run_results=results)
    payload = {"results": results, "quality_report": str(report_path)}
    print(json.dumps(payload, ensure_ascii=False, indent=2, default=str))
    # 閫鍑虹爜: 鏃犻噰闆嗗け璐ュ嵆瑙嗕负鎴愬姛(澧為噺妯″紡涓?鍏ㄩ儴璺宠繃"涔熷睘姝ｅ父鎴愬姛)
    return 0 if all(result.get("failed_count", 0) == 0 for result in results) else 1


if __name__ == "__main__":
    raise SystemExit(main())
